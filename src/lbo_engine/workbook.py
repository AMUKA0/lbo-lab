"""Export a deal as a LIVE Excel model.

Not a dump of the answers. Every calculated cell contains a formula, so an
analyst can click through the logic, change an assumption and watch the schedule
recalculate. That distinction is the whole point: a values-only export cannot be
audited, and an LBO model that cannot be audited is not trusted by anyone who
matters.

Conventions follow the ones every bank uses, because they are what make a
workbook readable by someone who did not build it:

    blue   hardcoded input      — the only cells you should type in
    black  formula, same sheet
    green  link to another sheet

Three things about this file are load-bearing:

**Iterative calculation is switched on in the file itself.** Interest is charged
on the average of opening and closing balances, closing depends on the sweep, and
the sweep depends on interest — the same circularity the Python engine solves by
iterating. Excel can do it too, but only with iteration enabled, and a workbook
that opens to a wall of circular-reference warnings reads as broken. So the flag
is written into the workbook, not left for the user to find.

**Inputs are named ranges, not cell addresses.** `Entry_EBITDA`, never
`Inputs!B7`. Analysts insert rows; named ranges survive that and addresses do not.

**The Checks sheet is not decoration.** Sources equal uses, the debt roll-forward
ties, cash never goes below the floor. A reviewer looks there first, and a model
that cannot show its own invariants holding is asking to be taken on trust.

What this does *not* yet reproduce is stated rather than approximated: mid-hold
capital events (recaps, divestitures, sponsor support) and the PIK-toggle
election are decisions the engine makes by search, and are refused here rather
than silently dropped — see `_reject_unsupported`.
"""

from __future__ import annotations

from dataclasses import dataclass

from lbo_engine.assumptions import Assumptions
from lbo_engine.sources_uses import build_sources_and_uses

# --- house style -------------------------------------------------------------
# Hex without the alpha prefix openpyxl also accepts, for legibility.
_BLUE = "0000CC"      # hardcoded input
_BLACK = "000000"     # formula, same sheet
_GREEN = "007A33"     # link to another sheet
_RULE = "BFBFBF"
_HEAD_BG = "1F3B2C"
_INPUT_BG = "EAF1FB"


@dataclass
class _Ref:
    """Where an input landed, so formulas can point at it by name."""

    name: str
    cell: str


def _allocate(pool: str, taken: list[str], capacity: str, negative: bool = True) -> str:
    """Formula for one claimant's share of a pool paid out in a fixed order.

    The same shape serves three jobs — the cash sweep, an asset sale's
    senior-first mandatory prepayment, and a sponsor repurchase working
    junior-first — and writing it once is the difference between three
    consistent formulas and three chances to fumble a MIN/MAX by hand.

    Each claimant takes what the ones ahead of it left, capped at what it can
    absorb.

    `taken` must be POSITIVE amounts, so each term needs an explicit `-` in
    front of the row it points at: repayment rows are stored negative, so they
    can be summed straight into a closing balance.

    That leading minus is load-bearing and its absence is silent. Subtracting a
    negative adds, so the second claimant sees a pool *larger* than the first
    one did and everything below the top of the waterfall over-repays. Nothing
    errors; the numbers are simply wrong. It is asserted rather than trusted
    because the bug is invisible in any structure with only one sweepable
    tranche — which was every fixture the suite had until HCA, with two, was
    exported.
    """
    for term in taken:
        assert term.startswith("-"), (
            f"_allocate needs the amount already taken as a positive: got {term!r}. "
            "Repayment rows are negative, so pass `-<cell>`, or the claimants "
            "below this one silently over-repay."
        )
    already = "+".join(taken) if taken else "0"
    sign = "-" if negative else ""
    return f"={sign}MIN(MAX({pool}-({already}),0),{capacity})"


def _coupon_ref(key: str, rate_rows: dict[str, int], column: str) -> str:
    """This year's coupon, whichever shape it was exported in.

    A flat coupon is a named scalar; a floating one is a row on Inputs. The
    Model sheet asks the same question either way, so adding a rate path did not
    require a second set of interest formulas — which is exactly the trap that
    makes a formula export drift from the engine.
    """
    if key in rate_rows:
        return f"Inputs!{column}{rate_rows[key]}"
    name = "Revolver_Rate" if key == "revolver" else f"{key}_Rate"
    return name


def _rate_row(inp, wb, row: int, text: str, values: list[float], name: str,
              years: int, label, Font, PatternFill, DefinedName) -> int:
    """A per-year coupon, as a blue input row with a named range over it."""
    label(inp, row, text, 1)
    for i in range(years):
        cell = inp.cell(row=row, column=2 + i, value=values[i] if i < len(values) else values[-1])
        cell.font = Font(color=_BLUE, bold=True)
        cell.fill = PatternFill("solid", fgColor=_INPUT_BG)
        cell.number_format = '0.00%'
    end = chr(ord("B") + years - 1)
    wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$B${row}:${end}${row}"))
    return row


def _money(ws, cell: str) -> None:
    ws[cell].number_format = '#,##0.0;(#,##0.0)'


def build_workbook(a: Assumptions, *, partial_from: Assumptions | None = None):
    """Return an openpyxl Workbook containing a live, formula-driven model.

    `partial_from` marks this as a schedule that stops at a break: `a` is the
    truncated deal, `partial_from` the full one it was cut from. Prefer
    `build_partial_workbook`, which does the truncation for you.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.workbook.defined_name import DefinedName

    # Run the engine first, always. It is cheap, it supplies the decisions the
    # event rows need, and — the reason it is unconditional — it refuses
    # structures that cannot finance. Without this the workbook would happily
    # print a schedule running a negative cash balance for a deal the model
    # itself declines to print, which is exactly the inconsistency the Checks
    # sheet exists to make impossible.
    solved = _solve_for_events(a)
    partial = partial_from is not None

    su = build_sources_and_uses(a)
    years = a.hold_years
    growth = a.growth_schedule()
    margin = a.margin_schedule()

    wb = Workbook()

    # THE flag. Without this the workbook opens complaining about circular
    # references and an analyst reasonably concludes the model is broken.
    # Only needed when interest is on average balances; harmless otherwise, but
    # setting it unconditionally would imply a circularity that isn't there.
    wb.calculation.iterate = a.interest_on_average_balance
    wb.calculation.iterateCount = 200
    wb.calculation.iterateDelta = 1e-10

    # slug → the Inputs row holding that coupon's per-year path, for coupons
    # that float. A flat coupon stays a single named cell and is absent here.
    # Declared up front because the tranche block fills it while walking the
    # stack and the Model sheet reads it long afterwards.
    rate_rows: dict[str, int] = {}

    thin = Side(style="thin", color=_RULE)

    def head(ws, row: int, text: str, span: int = 8) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        for i in range(1, span + 1):
            ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=_HEAD_BG)

    def label(ws, row: int, text: str, indent: int = 0) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.alignment = Alignment(indent=indent)
        c.font = Font(color=_BLACK)

    # ----------------------------------------------------------------- Inputs
    inp = wb.active
    inp.title = "Inputs"
    inp.column_dimensions["A"].width = 38
    for col in "BCDEFGHIJKLMNOP":
        inp.column_dimensions[col].width = 12

    refs: dict[str, _Ref] = {}

    def put(row: int, text: str, value, name: str, fmt: str | None = None, col: int = 2):
        """One blue input, registered as a named range."""
        label(inp, row, text, indent=1)
        c = inp.cell(row=row, column=col, value=value)
        c.font = Font(color=_BLUE, bold=True)
        c.fill = PatternFill("solid", fgColor=_INPUT_BG)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        if fmt:
            c.number_format = fmt
        ref = _Ref(name, f"Inputs!${c.column_letter}${row}")
        refs[name] = ref
        wb.defined_names.add(DefinedName(name, attr_text=ref.cell))
        return ref

    r = 1
    inp.cell(row=r, column=1, value="LBO model — inputs").font = Font(bold=True, size=14)
    r += 1
    inp.cell(row=r, column=1, value="Blue cells are inputs. Everything else is calculated.").font = Font(
        italic=True, color="666666", size=9
    )
    r += 1

    if partial:
        # Unmissable, and repeated on the Returns sheet. A partial model that
        # does not announce itself is the most dangerous file in this project:
        # every number on it is arithmetically right, and the conclusion someone
        # would draw from it is wrong.
        r += 1
        for line in (
            f"PARTIAL MODEL - this deal does not survive its "
            f"{partial_from.hold_years}-year hold.",
            f"The schedule stops at year {a.hold_years}, the last year the structure "
            f"could service itself. It breaks in year {a.hold_years + 1}. There is no "
            "exit and no return, because neither happened.",
        ):
            c = inp.cell(row=r, column=1, value=line)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            for i in range(1, 9):
                inp.cell(row=r, column=i).fill = PatternFill("solid", fgColor=_ALARM_BG)
            r += 1
    r += 1

    head(inp, r, "Entry"); r += 1
    put(r, "Entry EBITDA ($m)", a.entry_ebitda, "Entry_EBITDA", '#,##0.0'); r += 1
    put(r, "Entry multiple (× EBITDA)", a.entry_multiple, "Entry_Multiple", '0.00"x"'); r += 1
    put(r, "Entry revenue ($m)", a.operating.entry_revenue, "Entry_Revenue", '#,##0.0'); r += 2

    head(inp, r, "Operating"); r += 1
    put(r, "D&A (% of revenue)", a.operating.da_pct_revenue, "DA_Pct", '0.0%'); r += 1
    put(r, "Capex (% of revenue)", a.operating.capex_pct_revenue, "Capex_Pct", '0.0%'); r += 1
    put(r, "Working capital (% of revenue)", a.operating.nwc_pct_revenue, "NWC_Pct", '0.0%'); r += 1
    put(r, "Tax rate", a.operating.tax_rate, "Tax_Rate", '0.0%'); r += 1
    put(r, "NOL shelter limit (% of income)", a.nol_limit_pct, "NOL_Limit", '0.0%'); r += 1
    # §163(j), as three switches rather than one. An analyst asked to "run it
    # pre-2018" needs to turn the cap off; one asked to "run it on the old ATI
    # basis" needs the D&A add-back back. Both are one cell.
    lim = a.interest_limitation
    put(r, "§163(j) interest cap applies (1 = yes)",
        1 if lim.enabled else 0, "Interest_Limit_On", '0'); r += 1
    put(r, "§163(j) — deductible % of ATI", lim.pct_of_ati, "Interest_Limit_Pct", '0.0%'); r += 1
    put(r, "§163(j) — ATI adds back D&A (1 = pre-2022 basis)",
        1 if lim.ati_basis == "ebitda" else 0, "Interest_Limit_DA", '0'); r += 2

    head(inp, r, "Operating path — by year"); r += 1
    label(inp, r, "Year", indent=1)
    for i in range(years):
        c = inp.cell(row=r, column=2 + i, value=i + 1)
        c.font = Font(bold=True, color=_BLACK)
        c.alignment = Alignment(horizontal="center")
    r += 1
    growth_row = r
    label(inp, r, "Revenue growth", indent=1)
    for i, g in enumerate(growth):
        c = inp.cell(row=r, column=2 + i, value=g)
        c.font = Font(color=_BLUE, bold=True)
        c.fill = PatternFill("solid", fgColor=_INPUT_BG)
        c.number_format = '0.0%'
    r += 1
    margin_row = r
    label(inp, r, "EBITDA margin", indent=1)
    for i, m in enumerate(margin):
        c = inp.cell(row=r, column=2 + i, value=m)
        c.font = Font(color=_BLUE, bold=True)
        c.fill = PatternFill("solid", fgColor=_INPUT_BG)
        c.number_format = '0.0%'
    r += 2

    head(inp, r, "Capital structure"); r += 1
    tranche_rows: list[dict] = []
    for idx, t in enumerate(a.tranches):
        label(inp, r, t.name, indent=0)
        inp.cell(row=r, column=1).font = Font(bold=True)
        r += 1
        # Positional, not derived from the name — so renaming a tranche in the
        # workbook round-trips instead of breaking the lookup.
        slug = f"T{idx + 1}"
        rows = {
            "name": put(r, "Tranche name", t.name, f"{slug}_Name"),
        }
        r += 1
        rows["turns"] = put(r, "Leverage (turns of EBITDA)", t.leverage_turns,
                            f"{slug}_Turns", '0.00"x"')
        r += 1
        if isinstance(t.cash_rate, list):
            # A floating coupon. Written as a row so an analyst can shape a rate
            # path in the sheet — which is the whole point of exporting it: the
            # single most valuable thing to flex on a 2007 vintage is what
            # happened to LIBOR, and a scalar cell cannot express it.
            rate_rows[slug] = _rate_row(
                inp, wb, r, "Cash coupon (by year)", t.cash_rate, f"{slug}_Rate",
                years, label, Font, PatternFill, DefinedName)
            r += 1
        else:
            rows["cash"] = put(r, "Cash coupon", t.cash_rate, f"{slug}_Rate", '0.00%')
            r += 1
        rows["pik"] = put(r, "PIK rate", t.pik_rate, f"{slug}_PIK", '0.00%'); r += 1
        rows["amort"] = put(r, "Mandatory amortisation (% of original)", t.mandatory_amort_pct,
                            f"{slug}_Amort", '0.0%'); r += 1
        rows["sweep_flag"] = put(r, "Sweeps against this tranche (1/0)",
                                 1 if t.sweepable else 0, f"{slug}_Sweepable", '0')
        r += 1
        rows["toggle"] = t.pik_toggle
        rows["premium"] = t.pik_toggle_premium
        rows["sweepable"] = t.sweepable
        rows["name"] = t.name
        rows["slug"] = slug
        tranche_rows.append(rows)
        r += 1

    head(inp, r, "Revolver & cash policy"); r += 1
    put(r, "Revolver commitment ($m)", a.revolver.commitment, "Revolver_Commitment", '#,##0.0'); r += 1
    if isinstance(a.revolver.cash_rate, list):
        rate_rows["revolver"] = _rate_row(
            inp, wb, r, "Revolver rate (by year)", a.revolver.cash_rate,
            "Revolver_Rate", years, label, Font, PatternFill, DefinedName)
        r += 1
    else:
        put(r, "Revolver rate", a.revolver.cash_rate, "Revolver_Rate", '0.00%')
        r += 1
    put(r, "Undrawn commitment fee", a.revolver.undrawn_fee, "Undrawn_Fee", '0.00%'); r += 1
    put(r, "Minimum cash ($m)", a.minimum_cash, "Minimum_Cash", '#,##0.0'); r += 1
    put(r, "Deposit rate on cash", a.cash_deposit_rate, "Deposit_Rate", '0.00%'); r += 1
    # Carried so the round trip is lossless. The sheet reads the ELECTIONS, which
    # are exported below as overridable 1/0 inputs — this is the policy that
    # produced them, and losing it would silently change the next export.
    put(r, "PIK election headroom (% of revolver)", a.pik_election_headroom,
        "PIK_Headroom", '0%'); r += 1
    put(r, "Cash sweep (% of excess)", a.cash_sweep_pct, "Sweep_Pct", '0.0%'); r += 2

    # --- mid-hold capital events, one row per year so an analyst can move an
    # event to a different year or zero it out without touching a formula.
    ev: dict[str, int] = {}

    def year_row(text: str, values, name: str, fmt: str = '#,##0.0') -> None:
        nonlocal r
        label(inp, r, text, indent=1)
        for i in range(years):
            c = inp.cell(row=r, column=2 + i, value=values[i])
            c.font = Font(color=_BLUE, bold=True)
            c.fill = PatternFill("solid", fgColor=_INPUT_BG)
            c.number_format = fmt
        end = chr(ord("B") + years - 1)
        wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$B${r}:${end}${r}"))
        ev[name] = r
        r += 1

    def per_year(fn):
        return [float(fn(y)) for y in range(1, years + 1)]

    # Covenant levels as per-year inputs, not constants buried in a Checks
    # formula. An agreement steps down, and an analyst asked "what if we had
    # negotiated half a turn more headroom" must be able to answer it by typing
    # in a cell.
    cov_leverage = a.covenant_schedule("net_leverage_ceiling")
    cov_coverage = a.covenant_schedule("interest_coverage_floor")
    if cov_leverage is not None or cov_coverage is not None:
        head(inp, r, "Maintenance covenants — by year"); r += 1
        label(inp, r, "Year", indent=1)
        for i in range(years):
            c = inp.cell(row=r, column=2 + i, value=i + 1)
            c.font = Font(bold=True, color=_BLACK)
            c.alignment = Alignment(horizontal="center")
        r += 1
        if cov_leverage is not None:
            year_row("Net leverage ceiling (x)", cov_leverage,
                     "Leverage_Covenant", '0.00"x"')
        if cov_coverage is not None:
            year_row("Interest coverage floor (x)", cov_coverage,
                     "Coverage_Covenant", '0.00"x"')
        r += 1

    has_events = bool(a.divestitures or a.injections or a.recaps)
    toggles = [i for i, t in enumerate(a.tranches) if t.pik_toggle]

    # A recap sized by target leverage, and a toggle election, are decisions the
    # engine reaches by looking at the solved year. Exporting each as a number is
    # honest about its origin and leaves it flexible — an analyst can override
    # any of them and the schedule follows.
    if has_events or toggles:
        head(inp, r, "Capital events — by year"); r += 1
        label(inp, r, "Year", indent=1)
        for i in range(years):
            c = inp.cell(row=r, column=2 + i, value=i + 1)
            c.font = Font(bold=True, color=_BLACK)
            c.alignment = Alignment(horizontal="center")
        r += 1

    if a.divestitures:
        year_row("Divestiture proceeds, net of costs and tax",
                 per_year(lambda y: solved.years[y - 1].divestiture_proceeds),
                 "Divest_Net")
    if a.divestitures:
        # Revenue that leaves with a sold business. Without it the model books a
        # working-capital release from revenue that departed and was already
        # inside the sale consideration — the engine excludes it, so the
        # workbook must too or the two disagree.
        year_row("Revenue leaving with divested businesses",
                 per_year(lambda y: sum(d.revenue_removed for d in a.divestitures_for(y))),
                 "Divest_Revenue")
    if a.recaps:
        year_row("Recap financing fee",
                 per_year(lambda y: solved.years[y - 1].recap_fee),
                 "Recap_Fee")
    if a.injections:
        year_row("Sponsor cash injected",
                 per_year(lambda y: solved.years[y - 1].equity_injected),
                 "Injection_Cash")
        year_row("Debt extinguished by the sponsor",
                 per_year(lambda y: solved.years[y - 1].debt_retired),
                 "Injection_Retired")
    if a.recaps:
        year_row("Recap debt raised",
                 per_year(lambda y: solved.years[y - 1].recap_raised),
                 "Recap_Raised")
    for idx in toggles:
        year_row(f"{a.tranches[idx].name} — PIK elected (1/0)",
                 per_year(lambda y, n=a.tranches[idx].name:
                          1 if n in solved.years[y - 1].pik_elections else 0),
                 f"T{idx + 1}_Elected", '0')

    head(inp, r, "Fees & exit"); r += 1
    put(r, "Transaction fees (% of EV)", a.transaction_fee_pct_ev, "Txn_Fee_Pct", '0.00%'); r += 1
    put(r, "Financing fees (% of debt)", a.financing_fee_pct_debt, "Fin_Fee_Pct", '0.00%'); r += 1
    put(r, "Facility tenor (years)", a.financing_fee_tenor_years, "Fee_Tenor", '0'); r += 1
    put(r, "Exit costs (% of exit EV)", a.exit_fee_pct_ev, "Exit_Fee_Pct", '0.00%'); r += 1
    put(r, "Hold period (years)", a.hold_years, "Hold_Years", '0'); r += 1
    put(r, "Exit multiple (× EBITDA)", a.exit_multiple, "Exit_Multiple", '0.00"x"'); r += 1

    def event_ref(name: str) -> str:
        """A per-year event row, addressed by column."""
        return f"Inputs!${{c}}${ev[name]}"

    growth_ref = f"Inputs!${{c}}${growth_row}"
    margin_ref = f"Inputs!${{c}}${margin_row}"

    # The per-year rows as ranges, and the hold as a name, so a reader can
    # recover the whole operating path without knowing where anything sits.
    last_col = chr(ord("B") + years - 1)
    wb.defined_names.add(DefinedName(
        "Revenue_Growth", attr_text=f"Inputs!$B${growth_row}:${last_col}${growth_row}"))
    wb.defined_names.add(DefinedName(
        "EBITDA_Margin", attr_text=f"Inputs!$B${margin_row}:${last_col}${margin_row}"))

    # ------------------------------------------------------- Sources & Uses
    su_ws = wb.create_sheet("S&U")
    su_ws.column_dimensions["A"].width = 38
    su_ws.column_dimensions["B"].width = 14
    su_ws.cell(row=1, column=1, value="Sources & uses").font = Font(bold=True, size=14)

    r = 3
    head(su_ws, r, "Uses", 2); r += 1
    label(su_ws, r, "Purchase enterprise value", 1)
    su_ws.cell(row=r, column=2, value="=Entry_EBITDA*Entry_Multiple").font = Font(color=_GREEN)
    ev_cell = f"'S&U'!$B${r}"; _money(su_ws, f"B{r}"); r += 1
    label(su_ws, r, "Transaction fees", 1)
    su_ws.cell(row=r, column=2, value=f"={ev_cell}*Txn_Fee_Pct").font = Font(color=_BLACK)
    txn_cell = f"'S&U'!$B${r}"; _money(su_ws, f"B{r}"); r += 1
    fin_row = r
    label(su_ws, r, "Financing fees", 1)
    _money(su_ws, f"B{r}"); r += 1
    label(su_ws, r, "Cash to balance sheet", 1)
    su_ws.cell(row=r, column=2, value="=Minimum_Cash").font = Font(color=_GREEN)
    cash_cell = f"'S&U'!$B${r}"; _money(su_ws, f"B{r}"); r += 1
    uses_row = r
    label(su_ws, r, "Total uses", 0)
    su_ws.cell(row=r, column=1).font = Font(bold=True)
    su_ws.cell(row=r, column=2,
               value=f"={ev_cell}+{txn_cell}+'S&U'!$B${fin_row}+{cash_cell}").font = Font(bold=True)
    _money(su_ws, f"B{r}")
    uses_cell = f"'S&U'!$B${uses_row}"
    r += 2

    head(su_ws, r, "Sources", 2); r += 1
    tranche_amount_cells: list[str] = []
    for tr in tranche_rows:
        label(su_ws, r, tr["name"], 1)
        su_ws.cell(row=r, column=2, value=f"=Entry_EBITDA*{tr['slug']}_Turns").font = Font(color=_GREEN)
        _money(su_ws, f"B{r}")
        tranche_amount_cells.append(f"'S&U'!$B${r}")
        tr["amount_cell"] = f"'S&U'!$B${r}"
        r += 1
    debt_row = r
    label(su_ws, r, "Total debt", 0)
    su_ws.cell(row=r, column=1).font = Font(bold=True)
    su_ws.cell(row=r, column=2, value="=" + "+".join(tranche_amount_cells)).font = Font(bold=True)
    _money(su_ws, f"B{r}")
    debt_cell = f"'S&U'!$B${debt_row}"
    r += 1
    label(su_ws, r, "Sponsor equity (the plug)", 1)
    su_ws.cell(row=r, column=2, value=f"={uses_cell}-{debt_cell}").font = Font(color=_BLACK, bold=True)
    equity_cell = f"'S&U'!$B${r}"; _money(su_ws, f"B{r}"); r += 1
    label(su_ws, r, "Total sources", 0)
    su_ws.cell(row=r, column=1).font = Font(bold=True)
    su_ws.cell(row=r, column=2, value=f"={debt_cell}+{equity_cell}").font = Font(bold=True)
    _money(su_ws, f"B{r}")
    sources_cell = f"'S&U'!$B${r}"

    # Financing fees are a % of funded debt, which is only known once the
    # tranches are sized — so it is written after them and points back.
    su_ws.cell(row=fin_row, column=2, value=f"={debt_cell}*Fin_Fee_Pct").font = Font(color=_BLACK)
    fin_cell = f"'S&U'!$B${fin_row}"

    # ------------------------------------------------------------------ Model
    model = wb.create_sheet("Model")
    model.column_dimensions["A"].width = 44
    for i in range(years):
        model.column_dimensions[chr(ord("B") + i)].width = 13

    def col(i: int) -> str:
        return chr(ord("B") + i)

    m = 1
    model.cell(row=m, column=1, value="Annual schedule").font = Font(bold=True, size=14)
    m += 2
    label(model, m, "Year")
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=i + 1)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=thin)
    m += 1

    rows: dict[str, int] = {}

    def line(name: str, formula_for, bold: bool = False, fmt: str = '#,##0.0',
             indent: int = 1, colour: str = _BLACK) -> int:
        nonlocal m
        label(model, m, name, indent)
        if bold:
            model.cell(row=m, column=1).font = Font(bold=True)
        for i in range(years):
            c = model.cell(row=m, column=2 + i, value=formula_for(i))
            c.font = Font(color=colour, bold=bold)
            c.number_format = fmt
        rows[name] = m
        m += 1
        return m - 1

    def prior(name: str, i: int, fallback: str) -> str:
        """Previous year's cell, or a fallback for year one."""
        return fallback if i == 0 else f"{col(i - 1)}{rows[name]}"

    head(model, m, "Operating", years + 1); m += 1

    # Revenue is written directly rather than through `line`, because year one
    # points at an input and later years point at the cell to their left — a
    # self-referencing row the generic helper cannot express.
    label(model, m, "Revenue", 1)
    for i in range(years):
        base = "Entry_Revenue" if i == 0 else f"{col(i-1)}{m}"
        c = model.cell(row=m, column=2 + i, value=f"={base}*(1+{growth_ref.format(c=col(i))})")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    rows["Revenue"] = m; m += 1

    line("EBITDA", lambda i: f"={col(i)}{rows['Revenue']}*{margin_ref.format(c=col(i))}", bold=True)
    line("D&A", lambda i: f"=-{col(i)}{rows['Revenue']}*DA_Pct")
    line("EBIT", lambda i: f"={col(i)}{rows['EBITDA']}+{col(i)}{rows['D&A']}", bold=True)
    line("Capex", lambda i: f"=-{col(i)}{rows['Revenue']}*Capex_Pct")
    label(model, m, "Change in working capital", 1)
    for i in range(years):
        base = "Entry_Revenue" if i == 0 else f"{col(i-1)}{rows['Revenue']}"
        # A fall in revenue caused by selling a business is inorganic, so it
        # must not drive a working-capital release: that cash went with the
        # business. The sale completed at the END of the prior year, so it is
        # that year's figure which is added back here.
        organic = f"{col(i)}{rows['Revenue']}-{base}"
        if a.divestitures and i > 0:
            organic += f"+{event_ref('Divest_Revenue').format(c=col(i-1))}"
        c = model.cell(row=m, column=2 + i, value=f"=-NWC_Pct*({organic})")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    rows["dNWC"] = m; m += 1
    # Recap fees amortise over the same tenor, from the year after they are
    # incurred, exactly as the entry financing fee does.
    def fee_amort(i: int) -> str:
        if not a.recaps or i == 0:
            return f"=-{fin_cell}/Fee_Tenor"
        prior = "+".join(event_ref("Recap_Fee").format(c=col(k)) for k in range(i))
        return f"=-({fin_cell}+{prior})/Fee_Tenor"

    line("Financing fee amortisation", fee_amort)
    m += 1

    # --- debt, per tranche. Opening / PIK / interest / amortisation / closing.
    head(model, m, "Debt schedule", years + 1); m += 1
    # Senior-first, filled top down as the loop goes. The junior-first
    # retirement cannot be, so its rows are reserved and filled after the loop.
    divest_taken: list[int] = []
    for tr in tranche_rows:
        slug, nm = tr["slug"], tr["name"]
        label(model, m, nm, 0); model.cell(row=m, column=1).font = Font(bold=True); m += 1

        position = tranche_rows.index(tr)

        label(model, m, "Brought forward", 2)
        for i in range(years):
            v = f"={tr['amount_cell']}" if i == 0 else f"={col(i-1)}{{CLOSING}}"
            c = model.cell(row=m, column=2 + i, value=v)
            c.font = Font(color=_GREEN if i == 0 else _BLACK); c.number_format = '#,##0.0'
        tr["bf_row"] = m; m += 1

        if a.injections:
            # Junior-first. An elective repurchase buys the most discounted
            # paper, which is the junior end — unlike an asset sale, whose
            # proceeds are a contractual senior-first prepayment.
            # Reserved now, filled after the loop: junior-first means the
            # tranches below this one claim first, and their rows do not exist
            # yet. Excel does not care about write order, only references.
            label(model, m, "Less: extinguished by the sponsor", 2)
            tr["retire_row"] = m
            m += 1

        label(model, m, "Opening", 2)
        for i in range(years):
            parts = [f"{col(i)}{tr['bf_row']}"]
            if a.injections:
                parts.append(f"{col(i)}{tr['retire_row']}")
            c = model.cell(row=m, column=2 + i, value="=" + "+".join(parts))
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        tr["open_row"] = m; m += 1

        label(model, m, "PIK accrual", 2)
        for i in range(years):
            if tr["toggle"]:
                elected = event_ref(f"T{position + 1}_Elected").format(c=col(i))
                coupon = _coupon_ref(slug, rate_rows, col(i))
                rate = (f"IF({elected}=1,{slug}_PIK+{coupon}+{tr['premium']},{slug}_PIK)")
            else:
                rate = f"{slug}_PIK"
            c = model.cell(row=m, column=2 + i, value=f"={col(i)}{tr['open_row']}*{rate}")
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        tr["pik_row"] = m; m += 1

        label(model, m, "Mandatory amortisation", 2)
        for i in range(years):
            # Amortisation is a % of ORIGINAL principal, and any recap debt
            # raised in an earlier year joins that base — otherwise the
            # incremental debt never amortises again.
            base = tr["amount_cell"]
            if a.recaps and i > 0:
                prior = "+".join(f"{col(k)}{{RECAP}}" for k in range(i))
                base = f"({tr['amount_cell']}+{prior})"
            outstanding = f"{col(i)}{tr['open_row']}+{col(i)}{tr['pik_row']}"
            scheduled = f"MIN({base}*{slug}_Amort,{outstanding})"
            # A tranche reaching maturity owes its whole remaining balance, not
            # its scheduled slice — unless it is assumed to be refinanced, in
            # which case nothing falls due and it rolls at the stepped-up rate.
            if t.maturity_years == i + 1 and not t.refinance_at_maturity:
                scheduled = f"({outstanding})"
            c = model.cell(row=m, column=2 + i, value=f"=-{scheduled}")
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        tr["amort_row"] = m; m += 1

        label(model, m, "Cash sweep", 2)
        tr["sweep_row"] = m; m += 1   # filled once the waterfall exists

        # Two closings, because the engine has two. Interest is struck on the
        # balance after the waterfall but BEFORE year-end capital events — money
        # raised or repaid on 31 December earned no coupon that year.
        label(model, m, "Closing, before year-end events", 2)
        for i in range(years):
            c = model.cell(row=m, column=2 + i, value=(
                f"={col(i)}{tr['open_row']}+{col(i)}{tr['pik_row']}"
                f"+{col(i)}{tr['amort_row']}+{col(i)}{tr['sweep_row']}"))
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        tr["close_row"] = m; m += 1

        if a.divestitures:
            # Senior-first: this tranche absorbs only what the ones above it
            # left. `divest_taken` accumulates their repayment rows, and each is
            # negative, so it is subtracted to give a positive amount consumed.
            label(model, m, "Less: divestiture proceeds", 2)
            for i in range(years):
                # The revolver is the most senior claim and is cleared first,
                # so the tranches share only what is left of the proceeds.
                pool = (f"MAX({event_ref('Divest_Net').format(c=col(i))}"
                        f"-{col(i)}{{REVCLOSE}},0)")
                c = model.cell(row=m, column=2 + i, value=_allocate(
                    pool,
                    [f"-{col(i)}{x}" for x in divest_taken],
                    f"{col(i)}{tr['close_row']}"))
                c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
            tr["divest_row"] = m
            divest_taken.append(m)
            m += 1

        if a.recaps:
            label(model, m, "Recap debt raised", 2)
            target = (a.recaps[0].tranche or a.tranches[0].name)
            for i in range(years):
                v = event_ref("Recap_Raised").format(c=col(i)) if tr["name"] == target else "0"
                c = model.cell(row=m, column=2 + i, value=f"={v}")
                c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
            tr["recap_row"] = m; m += 1

        label(model, m, "Closing", 2)
        for i in range(years):
            parts = [f"{col(i)}{tr['close_row']}"]
            if a.divestitures:
                parts.append(f"{col(i)}{tr['divest_row']}")
            if a.recaps:
                parts.append(f"{col(i)}{tr['recap_row']}")
            c = model.cell(row=m, column=2 + i, value="=" + "+".join(parts))
            c.font = Font(color=_BLACK, bold=True); c.number_format = '#,##0.0'
        tr["final_row"] = m; m += 1

        label(model, m, "Cash interest", 2)
        for i in range(years):
            # The circularity, or the escape from it. On average balances this
            # cell depends on the closing balance, which depends on the sweep,
            # which depends on this cell — resolved by the iteration flag set on
            # the workbook. The opening-balance convention is acyclic, and is
            # what makes the exported workbook independently recalculable in a
            # test; see tests/test_workbook.py.
            basis = (f"AVERAGE({col(i)}{tr['open_row']},{col(i)}{tr['close_row']})"
                     if a.interest_on_average_balance else f"{col(i)}{tr['open_row']}")
            coupon = _coupon_ref(slug, rate_rows, col(i))
            c = model.cell(row=m, column=2 + i, value=f"={coupon}*{basis}")
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        tr["int_row"] = m; m += 2

        if a.recaps:
            for i in range(1, years):
                cell = model.cell(row=tr["amort_row"], column=2 + i)
                cell.value = str(cell.value).replace("{RECAP}", str(tr["recap_row"]))

        # The brought-forward row could not name the final closing row until it
        # existed, so it is patched now.
        for i in range(1, years):
            model.cell(row=tr["bf_row"], column=2 + i).value = (
                f"={col(i-1)}{tr['final_row']}")

    if a.injections:
        # Junior-first, so walk the stack from the bottom.
        taken_rows: list[int] = []
        for tr in reversed(tranche_rows):
            for i in range(years):
                c = model.cell(row=tr["retire_row"], column=2 + i, value=_allocate(
                    event_ref("Injection_Retired").format(c=col(i)),
                    [f"-{col(i)}{x}" for x in taken_rows],
                    f"{col(i)}{tr['bf_row']}"))
                c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
            taken_rows.append(tr["retire_row"])

    # --- revolver
    label(model, m, "Revolver", 0); model.cell(row=m, column=1).font = Font(bold=True); m += 1
    label(model, m, "Opening", 2)
    rev_open = m
    for i in range(years):
        # Filled once the final closing row exists — it moves depending on
        # whether there are divestitures to repay from.
        c = model.cell(row=m, column=2 + i, value=0)
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1
    label(model, m, "Draw", 2); rev_draw = m; m += 1
    label(model, m, "Repayment", 2); rev_repay = m; m += 1
    label(model, m, "Closing, before year-end events", 2)
    rev_close = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i,
                       value=f"={col(i)}{rev_open}+{col(i)}{rev_draw}-{col(i)}{rev_repay}")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    # Divestiture proceeds clear the revolver BEFORE the tranches — it is the
    # most senior claim, and a mandatory prepayment runs top down. The tranche
    # rows above already assume this happened: their pool is the proceeds less
    # the revolver's pre-event balance. Without this row the revolver kept that
    # balance anyway, so total debt double-counted it, and the workbook printed
    # more debt than the engine on any deal that sold something in a year it had
    # drawn. RJR, whose whole plan is disposals, was out by a billion.
    rev_final = rev_close
    if a.divestitures:
        label(model, m, "Less: divestiture proceeds", 2)
        rev_divest = m
        for i in range(years):
            c = model.cell(row=m, column=2 + i, value=(
                f"=-MIN({event_ref('Divest_Net').format(c=col(i))},{col(i)}{rev_close})"))
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        m += 1
        label(model, m, "Closing", 2)
        rev_final = m
        for i in range(years):
            c = model.cell(row=m, column=2 + i,
                           value=f"={col(i)}{rev_close}+{col(i)}{rev_divest}")
            c.font = Font(color=_BLACK, bold=True); c.number_format = '#,##0.0'
        m += 1

    for i in range(years):
        v = 0 if i == 0 else f"={col(i-1)}{rev_final}"
        model.cell(row=rev_open, column=2 + i, value=v)

    label(model, m, "Undrawn commitment fee", 2)
    rev_fee = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i,
                       value=f"=-Undrawn_Fee*MAX(Revolver_Commitment-{col(i)}{rev_open},0)")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1
    label(model, m, "Interest", 2)
    rev_int = m
    for i in range(years):
        basis = (f"AVERAGE({col(i)}{rev_open},{col(i)}{rev_close})"
                 if a.interest_on_average_balance else f"{col(i)}{rev_open}")
        coupon = _coupon_ref("revolver", rate_rows, col(i))
        c = model.cell(row=m, column=2 + i, value=f"={coupon}*{basis}")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 2

    # --- income statement below the debt, because interest comes from above
    head(model, m, "Earnings & tax", years + 1); m += 1
    int_sum = "+".join(f"{{c}}{tr['int_row']}" for tr in tranche_rows) + f"+{{c}}{rev_int}"
    pik_sum = "+".join(f"{{c}}{tr['pik_row']}" for tr in tranche_rows)

    label(model, m, "Cash interest", 1)
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value="=-(" + int_sum.replace("{c}", col(i)) + ")")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    rows["CashInt"] = m; m += 1

    label(model, m, "PIK accrual", 1)
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value="=-(" + pik_sum.replace("{c}", col(i)) + ")")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    rows["PIK"] = m; m += 1

    # Reserved, not written: the formula points at the cash rows, which live in
    # the waterfall below. Excel does not care about write order, only about
    # cycles — and on the average-balance convention this one IS a cycle
    # (income raises cash, cash raises income), resolved by the same iteration
    # that already resolves the debt.
    label(model, m, "Interest income on cash", 1)
    inc_row = m; m += 1

    line("Pre-tax income", lambda i: (
        f"={col(i)}{rows['EBIT']}+{col(i)}{rows['CashInt']}+{col(i)}{rows['PIK']}"
        f"+{col(i)}{rows['Financing fee amortisation']}+{col(i)}{rev_fee}"
        f"+{col(i)}{inc_row}"), bold=True)

    # --- §163(j), before the NOL, because that is the order of the statute.
    # The interest sits above as a negative number; the cap works in positives,
    # so it is negated once here and stays positive down the block.
    label(model, m, "Business interest (§163(j))", 1)
    bus_int = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=-({col(i)}{rows['CashInt']}+{col(i)}{rows['PIK']}"
            f"+{col(i)}{rows['Financing fee amortisation']})"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    # The undrawn commitment fee is added back because it is NOT business
    # interest — it is deducted in full and sits outside the cap entirely.
    label(model, m, "Adjusted taxable income", 1)
    ati = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"={col(i)}{rows['EBIT']}+{col(i)}{rev_fee}"
            f"-Interest_Limit_DA*{col(i)}{rows['D&A']}"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    label(model, m, "Disallowed interest brought forward", 1)
    dis_open = m
    for i in range(years):
        # Denied interest is treated as paid again next year, and never expires.
        v = "=0" if i == 0 else f"={col(i-1)}{m+3}"
        c = model.cell(row=m, column=2 + i, value=v)
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    label(model, m, "Deductible capacity", 1)
    capacity = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=IF(Interest_Limit_On=1,"
            f"{col(i)}{inc_row}+Interest_Limit_Pct*MAX({col(i)}{ati},0),"
            f"{col(i)}{bus_int}+{col(i)}{dis_open})"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    label(model, m, "Interest deducted", 1)
    deducted = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=MIN({col(i)}{bus_int}+{col(i)}{dis_open},{col(i)}{capacity})"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    label(model, m, "Disallowed interest carried forward", 1)
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"={col(i)}{bus_int}+{col(i)}{dis_open}-{col(i)}{deducted}"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    # What tax is charged on: book profit, plus back the interest tax refused.
    line("Taxable income", lambda i: (
        f"={col(i)}{rows['Pre-tax income']}+{col(i)}{bus_int}-{col(i)}{deducted}"))
    taxable = rows["Taxable income"]

    label(model, m, "NOL brought forward", 1)
    nol_open = m
    for i in range(years):
        v = "=0" if i == 0 else f"={col(i-1)}{m+2}"
        c = model.cell(row=m, column=2 + i, value=v)
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1
    label(model, m, "NOL used", 1)
    nol_used = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=IF({col(i)}{taxable}>0,"
            f"MIN({col(i)}{nol_open},NOL_Limit*{col(i)}{taxable}),0)"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1
    label(model, m, "NOL carried forward", 1)
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=IF({col(i)}{taxable}>0,{col(i)}{nol_open}-{col(i)}{nol_used},"
            f"{col(i)}{nol_open}-{col(i)}{taxable})"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    line("Tax", lambda i: (
        f"=-IF({col(i)}{taxable}>0,"
        f"({col(i)}{taxable}-{col(i)}{nol_used})*Tax_Rate,0)"))
    line("Net income", lambda i: f"={col(i)}{rows['Pre-tax income']}+{col(i)}{rows['Tax']}", bold=True)
    m += 1

    head(model, m, "Cash flow & waterfall", years + 1); m += 1
    line("Cash available for debt service", lambda i: (
        f"={col(i)}{rows['Net income']}-{col(i)}{rows['D&A']}"
        f"-{col(i)}{rows['Financing fee amortisation']}-{col(i)}{rows['PIK']}"
        f"+{col(i)}{rows['Capex']}+{col(i)}{rows['dNWC']}"), bold=True)

    label(model, m, "Opening cash", 1)
    cash_open = m
    m += 1

    mand_sum = "+".join(f"{{c}}{tr['amort_row']}" for tr in tranche_rows)
    label(model, m, "Cash after mandatory amortisation", 1)
    after_mand = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"={col(i)}{cash_open}+{col(i)}{rows['Cash available for debt service']}"
            f"-Minimum_Cash+(" + mand_sum.replace("{c}", col(i)) + ")"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    for i in range(years):
        c = model.cell(row=rev_draw, column=2 + i, value=f"=MAX(-{col(i)}{after_mand},0)")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        c2 = model.cell(row=rev_repay, column=2 + i,
                        value=f"=MIN({col(i)}{rev_open},MAX({col(i)}{after_mand},0))")
        c2.font = Font(color=_BLACK); c2.number_format = '#,##0.0'

    label(model, m, "Cash for sweep", 1)
    sweep_pool = m
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=Sweep_Pct*MAX({col(i)}{after_mand}-{col(i)}{rev_repay},0)"))
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
    m += 1

    # Senior-first allocation: each tranche takes what is left after the ones
    # above it, capped at its own balance.
    # Every sweepable tranche gets the same formula, gated on its OWN blue
    # `{slug}_Sweepable` cell rather than on a flag baked in at export time.
    #
    # It used to be baked in: a non-sweepable tranche got a literal `=0`, and the
    # blue input sat there driving nothing. Flipping it changed no number. A blue
    # cell that does not drive anything is the commonest cause of a wrong answer
    # in a flexed model, and a sheet whose header says "blue cells are inputs"
    # actively invites the mistake.
    #
    # Because the flag is now live, EVERY tranche has to appear in `taken` —
    # what a junior tranche can absorb depends on what the seniors took, and that
    # is a decision the sheet makes at calculation time, not one we can make here.
    taken: list[str] = []
    for tr in tranche_rows:
        slug = tr["slug"]
        for i in range(years):
            # The sweep rows are negative, like every other repayment, so they
            # are negated here to give the positive amount already consumed.
            # Without that the subtraction ADDS, and the second sweepable
            # tranche sees a bigger pool than the first — which is exactly what
            # it did, undetected, until a case with two sweepable tranches was
            # exported.
            allocation = _allocate(
                f"{col(i)}{sweep_pool}",
                [f"-{col(i)}{t}" for t in taken],
                f"{col(i)}{tr['open_row']}+{col(i)}{tr['pik_row']}+{col(i)}{tr['amort_row']}")
            formula = f"=IF({slug}_Sweepable=1,{allocation.lstrip('=')},0)"
            c = model.cell(row=tr["sweep_row"], column=2 + i, value=formula)
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0'
        taken.append(str(tr["sweep_row"]))

    swept = ("+".join(f"{{c}}{t}" for t in taken)) if taken else "0"
    label(model, m, "Closing cash", 1)
    for i in range(years):
        c = model.cell(row=m, column=2 + i, value=(
            f"=Minimum_Cash+MAX({col(i)}{after_mand}-{col(i)}{rev_repay},0)"
            f"+(" + swept.replace("{c}", col(i)) + ")"))
        c.font = Font(color=_BLACK, bold=True); c.number_format = '#,##0.0'
    rows["Closing cash"] = m
    cash_close = m
    m += 1

    # Now that the closing row exists, point each year's opening cash at the
    # prior year's close.
    for i in range(years):
        base = f"{cash_cell}" if i == 0 else f"{col(i-1)}{cash_close}"
        # Rescue capital funds the year it goes into, so it arrives at the start.
        v = (f"={base}+{event_ref('Injection_Cash').format(c=col(i))}"
             if a.injections else f"={base}")
        c = model.cell(row=cash_open, column=2 + i, value=v)
        c.font = Font(color=_GREEN if i == 0 else _BLACK); c.number_format = '#,##0.0'

    # And fill the interest-income row reserved up in the income statement,
    # on the same balance convention the debt uses.
    for i in range(years):
        basis = (f"AVERAGE({col(i)}{cash_open},{col(i)}{cash_close})"
                 if a.interest_on_average_balance else f"{col(i)}{cash_open}")
        c = model.cell(row=inc_row, column=2 + i, value=f"=Deposit_Rate*{basis}")
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0'

    if a.divestitures:
        for tr in tranche_rows:
            for i in range(years):
                cell = model.cell(row=tr["divest_row"], column=2 + i)
                cell.value = str(cell.value).replace("{REVCLOSE}", str(rev_close))

    close_sum = "+".join(f"{{c}}{tr['final_row']}" for tr in tranche_rows) + f"+{{c}}{rev_final}"
    line("Total debt", lambda i: "=" + close_sum.replace("{c}", col(i)), bold=True)
    line("Net debt", lambda i: f"={col(i)}{rows['Total debt']}-{col(i)}{cash_close}", bold=True)

    # ---------------------------------------------------------------- Returns
    ret = wb.create_sheet("Returns")
    ret.column_dimensions["A"].width = 40
    ret.column_dimensions["B"].width = 16
    ret.cell(row=1, column=1, value="Exit & returns").font = Font(bold=True, size=14)
    last = col(years - 1)

    # Wide enough for a flow per year plus the close.
    for i in range(years + 1):
        ret.column_dimensions[chr(ord("B") + i)].width = 13

    rr = 3
    head(ret, rr, "Exit", years + 2); rr += 1
    for text, formula, fmt in (
        ("Terminal EBITDA", f"=Model!{last}{rows['EBITDA']}", '#,##0.0'),
        ("Exit enterprise value", f"=Model!{last}{rows['EBITDA']}*Exit_Multiple", '#,##0.0'),
        ("Less: net debt", f"=-Model!{last}{rows['Net debt']}", '#,##0.0'),
        ("Less: sale costs", f"=-Model!{last}{rows['EBITDA']}*Exit_Multiple*Exit_Fee_Pct", '#,##0.0'),
    ):
        label(ret, rr, text, 1)
        c = ret.cell(row=rr, column=2, value=formula)
        c.font = Font(color=_GREEN); c.number_format = fmt
        rr += 1
    label(ret, rr, "Exit equity", 0); ret.cell(row=rr, column=1).font = Font(bold=True)
    ret.cell(row=rr, column=2, value=f"=MAX(B{rr-3}+B{rr-2}+B{rr-1},0)").font = Font(bold=True)
    ret.cell(row=rr, column=2).number_format = '#,##0.0'
    exit_eq = f"Returns!$B${rr}"
    rr += 2

    # --- the sponsor's actual cash flows, year by year ----------------------
    #
    # This block exists because its absence was a real, shipped error. MOIC was
    # exit equity over the closing cheque and IRR was that multiple compounded
    # over the hold — both correct ONLY for a deal with no interim flows, and
    # asserted as correct for every deal. On HCA, whose sponsors took a $4.3bn
    # recap dividend in year four, the workbook reported 2.27x against the
    # engine's 2.95x. The debt that funded the dividend was in the model; the
    # dividend was not. On Hilton, where the sponsor injected rescue capital,
    # the same omission ran the other way and flattered the multiple.
    #
    # Laid out horizontally, year 0 to the exit, because that is the shape IRR
    # takes and because a reader can see the money move.
    head(ret, rr, "Sponsor cash flows", years + 2); rr += 1
    label(ret, rr, "Year", 1)
    for i in range(years + 1):
        c = ret.cell(row=rr, column=2 + i, value=i)
        c.font = Font(bold=True, color=_BLACK); c.alignment = Alignment(horizontal="center")
    rr += 1

    def flow_row(text: str, filler) -> int:
        nonlocal rr
        label(ret, rr, text, 1)
        for i in range(years + 1):
            c = ret.cell(row=rr, column=2 + i, value=filler(i))
            c.font = Font(color=_BLACK); c.number_format = '#,##0.0;(#,##0.0)'
        rr += 1
        return rr - 1

    fc = lambda i: chr(ord("B") + i)  # noqa: E731 - flow column for year i

    invested_row = flow_row(
        "Equity invested at close",
        lambda i: f"=-{equity_cell}" if i == 0 else "=0",
    )

    # An injection funds the year it goes into, so it is spent at that year's
    # START — index year-1 in the flow vector. Discounting it a year later than
    # it was paid would flatter the IRR of every rescued deal.
    if a.injections:
        injection_row = flow_row(
            "Follow-on equity",
            lambda i: (f"=-{event_ref('Injection_Cash').format(c=col(i))}"
                       if i < years else "=0"),
        )
    else:
        injection_row = None

    # A recap is a year-END event, so it lands at index `year`. What reaches the
    # sponsor is the debt raised less the financing fee on it.
    if a.recaps:
        dividend_row = flow_row(
            "Recap dividends",
            lambda i: (f"={event_ref('Recap_Raised').format(c=col(i - 1))}"
                       f"-{event_ref('Recap_Fee').format(c=col(i - 1))}"
                       if i >= 1 else "=0"),
        )
    else:
        dividend_row = None

    exit_row = flow_row(
        "Exit proceeds",
        lambda i: f"={exit_eq}" if i == years else "=0",
    )

    parts = [invested_row, exit_row] + [r for r in (injection_row, dividend_row) if r]
    label(ret, rr, "Net cash flow", 0); ret.cell(row=rr, column=1).font = Font(bold=True)
    for i in range(years + 1):
        formula = "=" + "+".join(f"{fc(i)}{r}" for r in sorted(parts))
        c = ret.cell(row=rr, column=2 + i, value=formula)
        c.font = Font(bold=True, color=_BLACK); c.number_format = '#,##0.0;(#,##0.0)'
    flow_row_index = rr
    flow_range = f"Returns!$B${rr}:${fc(years)}${rr}"
    rr += 2

    # --- returns, struck on those flows ------------------------------------
    head(ret, rr, "Returns", years + 2); rr += 1

    invested_terms = [f"{fc(i)}{invested_row}" for i in range(years + 1)]
    if injection_row:
        invested_terms += [f"{fc(i)}{injection_row}" for i in range(years + 1)]
    proceeds_terms = [f"{fc(i)}{exit_row}" for i in range(years + 1)]
    if dividend_row:
        proceeds_terms += [f"{fc(i)}{dividend_row}" for i in range(years + 1)]

    label(ret, rr, "Total invested", 1)
    ret.cell(row=rr, column=2, value="=-(" + "+".join(invested_terms) + ")").font = Font(color=_BLACK)
    ret.cell(row=rr, column=2).number_format = '#,##0.0'
    total_invested = f"Returns!$B${rr}"; rr += 1

    label(ret, rr, "Total proceeds", 1)
    ret.cell(row=rr, column=2, value="=" + "+".join(proceeds_terms)).font = Font(color=_BLACK)
    ret.cell(row=rr, column=2).number_format = '#,##0.0'
    total_proceeds = f"Returns!$B${rr}"; rr += 1

    label(ret, rr, "MOIC", 1)
    moic_cell = f"Returns!$B${rr}"
    ret.cell(row=rr, column=2, value=f"={total_proceeds}/{total_invested}").font = Font(bold=True)
    ret.cell(row=rr, column=2).number_format = '0.00"x"'
    ret.cell(row=rr, column=3, value="Every dollar back over every dollar in — "
             "rescue capital raises the denominator, recap dividends the numerator.").font = Font(
        italic=True, color="666666", size=9)
    rr += 1

    label(ret, rr, "IRR", 1)
    # IRR over the flow vector, not the compounded multiple. Those are the same
    # number only when nothing happens between close and exit, and the previous
    # version asserted the equivalence in a comment on files where it was false.
    ret.cell(row=rr, column=2, value=f"=IRR({flow_range})").font = Font(bold=True)
    ret.cell(row=rr, column=2).number_format = '0.0%'
    ret.cell(row=rr, column=3, value="Solved over the cash-flow row above. Annual "
             "periods, so IRR() rather than XIRR().").font = Font(
        italic=True, color="666666", size=9)
    rr += 2

    head(ret, rr, "Value creation bridge", years + 2); rr += 1
    bridge = [
        ("EBITDA growth × entry multiple",
         f"=(Model!{last}{rows['EBITDA']}-Entry_EBITDA)*Entry_Multiple"),
        ("Multiple change × exit EBITDA",
         f"=(Exit_Multiple-Entry_Multiple)*Model!{last}{rows['EBITDA']}"),
    ]

    # Divestiture proceeds are split OUT of deleveraging: paying debt down with
    # the proceeds of a sale is not the business earning its way down, and
    # merging the two is the commonest way a bridge flatters an operator.
    divested = ("+".join(event_ref("Divest_Net").format(c=col(i)) for i in range(years))
                if a.divestitures else "0")
    bridge.append((
        "Net debt paydown (from operations)",
        f"=({debt_cell}-{cash_cell})-Model!{last}{rows['Net debt']}-({divested})",
    ))
    if a.divestitures:
        bridge.append(("Divestiture proceeds", f"={divested}"))
    if a.recaps:
        # GROSS debt raised, not the net dividend: the fee is picked up in the
        # fee line below, and netting it here would leave the identity short.
        raised = "+".join(event_ref("Recap_Raised").format(c=col(i)) for i in range(years))
        bridge.append(("Recapitalisation", f"={raised}"))
    if a.injections:
        injected = "+".join(event_ref("Injection_Cash").format(c=col(i)) for i in range(years))
        bridge.append(("Follow-on equity", f"=-({injected})"))

    recap_fees = ("+".join(event_ref("Recap_Fee").format(c=col(i)) for i in range(years))
                  if a.recaps else "0")
    bridge.append((
        "Fees",
        f"=-({txn_cell}+{fin_cell}+({recap_fees})"
        f"+Model!{last}{rows['EBITDA']}*Exit_Multiple*Exit_Fee_Pct)",
    ))

    first_bridge = rr
    for text, formula in bridge:
        label(ret, rr, text, 1)
        c = ret.cell(row=rr, column=2, value=formula)
        c.font = Font(color=_BLACK); c.number_format = '#,##0.0;(#,##0.0)'
        rr += 1
    label(ret, rr, "Total value created", 0); ret.cell(row=rr, column=1).font = Font(bold=True)
    ret.cell(row=rr, column=2, value=f"=SUM(B{first_bridge}:B{rr-1})").font = Font(bold=True)
    ret.cell(row=rr, column=2).number_format = '#,##0.0;(#,##0.0)'
    bridge_total = f"Returns!$B${rr}"

    # ----------------------------------------------------------------- Checks
    chk = wb.create_sheet("Checks")
    chk.column_dimensions["A"].width = 46
    chk.column_dimensions["B"].width = 16
    chk.column_dimensions["C"].width = 12
    chk.cell(row=1, column=1, value="Checks").font = Font(bold=True, size=14)
    chk.cell(row=2, column=1,
             value="Every one must read OK. A model that cannot show its own "
                   "invariants holding is asking to be taken on trust.").font = Font(
        italic=True, color="666666", size=9)

    # Two kinds of check, and conflating them is a real mistake: an identity
    # must equal zero, whereas a limit only has to be the right side of zero.
    # Testing "cash never below the minimum" for equality would fail on every
    # healthy deal, which is how a Checks sheet trains people to ignore it.
    checks = [
        ("Sources equal uses", f"={sources_cell}-{uses_cell}", "zero"),
        ("Bridge reconciles to the equity gain",
         f"={bridge_total}-({total_proceeds}-{total_invested})", "zero"),
        # The flow vector and the multiple must describe the same deal. They
        # are computed from the same rows, so this is cheap — and it is exactly
        # the tie that was missing when MOIC ignored every interim flow while
        # the schedule above it modelled them.
        ("Net cash flows sum to proceeds less invested",
         f"=SUM({flow_range})-({total_proceeds}-{total_invested})", "zero"),
        ("Headroom over the minimum cash balance",
         f"=MIN(Model!B{cash_close}:{last}{cash_close})-Minimum_Cash", "positive"),
        ("Unused revolver commitment",
         f"=Revolver_Commitment-MAX(Model!B{rev_final}:{last}{rev_final})", "positive"),
    ]

    # Covenant headroom, one row per test, and only where the credit agreement
    # actually has one. A cov-lite structure gets no row rather than a row
    # reading "n/a" — the absence is the fact worth showing.
    leverage_ceiling = a.covenant_schedule("net_leverage_ceiling")
    coverage_floor = a.covenant_schedule("interest_coverage_floor")
    def worst_headroom(per_year: list[str]) -> str:
        """The tightest year's headroom. MIN across the hold, because a covenant
        that holds in four years out of five is a covenant that was breached."""
        return "=MIN(" + ",".join(per_year) + ")"

    if leverage_ceiling is not None:
        # Headroom in TURNS, per year, against that year's own ceiling — the
        # levels step down, so a single ceiling would test the wrong number in
        # every year but one.
        checks.append((
            "Leverage covenant headroom, tightest year (turns)",
            worst_headroom([
                f"Inputs!{col(i)}{ev['Leverage_Covenant']}"
                f"-Model!{col(i)}{rows['Net debt']}/Model!{col(i)}{rows['EBITDA']}"
                for i in range(years)
            ]),
            "positive"))

    if coverage_floor is not None:
        # Cash interest sits in the sheet as a negative, hence the sign flip.
        checks.append((
            "Coverage covenant headroom, tightest year (turns)",
            worst_headroom([
                f"Model!{col(i)}{rows['EBITDA']}/-Model!{col(i)}{rows['CashInt']}"
                f"-Inputs!{col(i)}{ev['Coverage_Covenant']}"
                for i in range(years)
            ]),
            "positive"))
    chk.cell(row=3, column=2, value="Value").font = Font(bold=True, size=9)
    chk.cell(row=3, column=3, value="Test").font = Font(bold=True, size=9)
    chk.cell(row=3, column=4, value="Result").font = Font(bold=True, size=9)
    cr = 4
    for text, formula, kind in checks:
        label(chk, cr, text, 1)
        c = chk.cell(row=cr, column=2, value=formula)
        c.font = Font(color=_BLACK); c.number_format = '#,##0.000;(#,##0.000)'
        chk.cell(row=cr, column=3,
                 value="= 0" if kind == "zero" else ">= 0").font = Font(size=9, color="666666")
        test = (f'=IF(ABS(B{cr})<0.01,"OK","CHECK")' if kind == "zero"
                else f'=IF(B{cr}>=-0.01,"OK","CHECK")')
        v = chk.cell(row=cr, column=4, value=test)
        v.font = Font(bold=True)
        cr += 1
    chk.column_dimensions["D"].width = 12

    if partial:
        _refuse_to_invent_an_exit(wb, ret, chk, a, partial_from, Alignment, Font)

    _freeze(inp, su_ws, model, chk)
    _prepare_for_print(wb, model, years, a)
    return wb


_ALARM_BG = "B03A2E"


def _refuse_to_invent_an_exit(wb, ret, chk, a, full, Alignment, Font) -> None:
    """Strip the Returns sheet of numbers, and say why.

    This is the one place a partial workbook must refuse rather than compute.
    Every other sheet is simply shorter; this one would be *wrong*. The sponsor
    did not sell in the final modelled year - the model stops there - so an exit
    value, a MOIC and an IRR struck on that balance sheet are not conservative
    estimates. They are arithmetically correct answers to a question nobody
    asked, and they are exactly the numbers a reader would quote.

    The sheet is rebuilt rather than deleted, because a missing sheet reads as an
    export bug and invites someone to compute the number by hand.
    """
    position = wb.sheetnames.index("Returns")
    wb.remove(ret)
    fresh = wb.create_sheet("Returns", position)
    fresh.column_dimensions["A"].width = 96
    fresh.cell(row=1, column=1, value="Exit & returns").font = Font(bold=True, size=14)

    lines = [
        ("No exit, and therefore no return.", True),
        ("", False),
        (f"This deal was underwritten on a {full.hold_years}-year hold and does not "
         f"survive it. The Model sheet runs to year {a.hold_years} - the last year "
         f"the structure could service itself - and it breaks in year "
         f"{a.hold_years + 1}.", False),
        ("", False),
        ("An exit value, a MOIC and an IRR could all be computed from the closing "
         "balance sheet on the Model sheet. They are deliberately absent. There was "
         "no sale in that year, so they would be arithmetically correct answers to a "
         "question nobody asked - and they are precisely the numbers a reader would "
         "quote.", False),
        ("", False),
        ("What IS here is real: every year up to the break, with the liquidity "
         "draining year by year, and the Checks sheet still proving the schedule "
         "ties. That is the useful output of a deal that did not work.", False),
    ]
    row = 3
    for text, strong in lines:
        c = fresh.cell(row=row, column=1, value=text)
        c.font = Font(bold=strong, size=12 if strong else 10,
                      color=_ALARM_BG if strong else "333333")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if text:
            fresh.row_dimensions[row].height = 15 * (1 + len(text) // 100)
        row += 1

    # The bridge check reconciles to an equity gain that no longer exists. Drop
    # it rather than leave a row reading #REF!, which would train a reader to
    # ignore the Checks sheet - the opposite of what it is for.
    for r in range(chk.max_row, 3, -1):
        if "Returns!" in str(chk.cell(row=r, column=2).value or ""):
            chk.delete_rows(r)


def _prepare_for_print(wb, model, years: int, a: Assumptions) -> None:
    """Page setup on every sheet.

    An IC exhibit gets printed. This is a small amount of work and its absence
    is loud: the first time someone hits Ctrl-P on a model with fourteen year
    columns and gets column H alone on page four, the file stops being credible
    regardless of what the numbers say.

    Four decisions, each with a reason:

    * **Landscape and fit-to-one-page-wide for the schedule**, portrait for the
      narrow sheets. Nothing here is worth splitting a year across a page break.
    * **Repeating title rows.** A schedule that runs past one page is unreadable
      without its year headers, and scrolling back is not an option on paper.
    * **A printed footer carrying the sheet name and page numbers**, because a
      loose page from a model with five sheets is otherwise unidentifiable.
    * **Print areas set explicitly**, so a stray value in a far column — an
      analyst's scratch note, which is exactly what people do — cannot silently
      drag a blank page into the printout.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    for sheet in wb.worksheets:
        wide = sheet is model and years > 4
        sheet.page_setup.orientation = "landscape" if wide else "portrait"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        # Zero means "as many pages as it takes" vertically, which is what a
        # long schedule needs — the alternative shrinks the type until nobody
        # can read it, which is the classic way this gets done badly.
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        sheet.page_margins.left = sheet.page_margins.right = 0.4
        sheet.page_margins.top = 0.6
        sheet.page_margins.bottom = 0.5

        # &L / &R are Excel's own left/right footer codes.
        sheet.oddFooter.left.text = f"{sheet.title} — LBO model"
        sheet.oddFooter.left.size = 8
        sheet.oddFooter.right.text = "Page &[Page] of &[Pages]"
        sheet.oddFooter.right.size = 8

        last_col = get_column_letter(max(sheet.max_column, 2))
        sheet.print_area = f"A1:{last_col}{sheet.max_row}"

    # The Model sheet is the one that runs long. Its header block is rows 1-3;
    # repeating them keeps the years legible on every page.
    model.print_title_rows = "1:3"


def _freeze(*sheets) -> None:
    for ws in sheets:
        ws.freeze_panes = "B4"
        ws.sheet_view.showGridLines = False


def _slug(name: str) -> str:
    """A defined-name-safe token. Excel names cannot contain spaces or most
    punctuation, and cannot collide with a cell reference."""
    out = "".join(ch if ch.isalnum() else "_" for ch in name)
    while "__" in out:
        out = out.replace("__", "_")
    return "T_" + out.strip("_")


def _solve_for_events(a: Assumptions):
    """Run the engine once, to read off the decisions it makes by search.

    A recap sized by target leverage resolves to a quantum only once the year
    is solved, and a PIK election is chosen by trying options. Both are exported
    as plain inputs: honest about their origin, and flexible, since an analyst
    can override either and watch the schedule follow.
    """
    from lbo_engine.engine import run_lbo

    try:
        return run_lbo(a)
    except ValueError as exc:
        # A structure that runs out of liquidity has no schedule to export. Say
        # that plainly rather than letting the engine's simulator-facing advice
        # surface as though the export were at fault.
        raise ValueError(
            "This structure runs out of liquidity during the hold, so there is "
            f"no complete schedule to export. The engine reports: {exc}"
        ) from exc


def _reject_unsupported(a: Assumptions) -> None:
    """Nothing is refused any more, and the reason the earlier version did is
    worth recording: it claimed these were decisions the engine makes by search
    rather than by formula. That was only true of one of them. Divestitures and
    sponsor support are fully specified inputs; a recap sized by target leverage
    is MAX(target x EBITDA - net debt, 0), which is arithmetic. Only the PIK
    election is a search, and exporting the engine's choice as an input the
    analyst can override handles it honestly.

    Kept as a hook so a future mechanic has an obvious place to declare itself
    unsupported rather than silently producing a workbook that disagrees.
    """
    return None


def build_partial_workbook(a: Assumptions):
    """A workbook for a deal that does not survive its hold.

    Exports the schedule up to the break, which is what the web page already
    shows and what a credit committee would actually want: the drain made
    legible, rather than the word "failed".

    Raises if the deal cannot service even one year, since there is then
    genuinely nothing to show.
    """
    from lbo_engine.partial import survivable_years, truncate

    survived = survivable_years(a)
    if survived <= 0:
        raise ValueError(
            "This structure cannot service even its first year, so there is no "
            "schedule to export. The problem is the structure, not the export."
        )
    return build_workbook(truncate(a, survived), partial_from=a)


def workbook_bytes(a: Assumptions, *, allow_partial: bool = False) -> bytes:
    """The workbook as bytes, for an HTTP response.

    `allow_partial` falls back to the schedule-up-to-the-break form rather than
    refusing outright. Off by default: a caller asking for "the model" should get
    an error rather than a silently shortened file.
    """
    import io

    try:
        wb = build_workbook(a)
    except ValueError:
        if not allow_partial:
            raise
        wb = build_partial_workbook(a)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
