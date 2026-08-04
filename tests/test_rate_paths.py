"""Coupons that move over the hold.

A review finding, and the largest genuine modelling gap in the project. Every
case in the library is a 2006–07 vintage carried through the collapse of LIBOR
to zero, and the paper was largely floating — Hilton's CMBS and mezzanine stack
and Dollar General's L+275 term loan both repriced down by 400–500bp within
eighteen months of close. The model charged the 2007 coupon for the whole hold
and filed that under "neutral" in its limitations list.

It is not neutral. On Hilton's realised column 200bp is worth roughly 0.46× of
MOIC — 15% of the answer, and the same order as the caveat that page flags most
prominently. `cash_rate` now accepts a per-year path exactly as revenue growth
and margin already do.
"""

import pytest

from lbo_engine import Assumptions, run_lbo
from lbo_engine.returns import sponsor_irr


def _path(deal, rates):
    d = deal.model_copy(deep=True)
    d.tranches[0].cash_rate = list(rates)
    return Assumptions.model_validate(d.model_dump())


class TestAScalarStillMeansFlat:
    def test_nothing_moves_for_an_unchanged_deal(self, rich_deal):
        """The path is additive. An existing deal's schedule must not shift."""
        assert isinstance(rich_deal.tranches[0].cash_rate, float)
        assert rich_deal.rate_schedule(rich_deal.tranches[0]) == [0.055] * 5

    def test_a_flat_path_equals_the_scalar(self, rich_deal):
        flat = _path(rich_deal, [rich_deal.tranches[0].cash_rate] * rich_deal.hold_years)
        base, spread = run_lbo(rich_deal), run_lbo(flat)
        for a, b in zip(base.years, spread.years):
            assert a.cash_interest_total == pytest.approx(b.cash_interest_total, abs=1e-9)


class TestThePathIsActuallyUsed:
    def test_each_year_charges_its_own_coupon(self, rich_deal):
        rates = [0.02, 0.04, 0.06, 0.08, 0.10]
        r = run_lbo(_path(rich_deal, rates))
        name = rich_deal.tranches[0].name
        for i, year in enumerate(r.years):
            t = year.tranches[name]
            basis = 0.5 * (t.opening + t.closing)  # average-balance convention
            assert t.cash_interest == pytest.approx(rates[i] * basis, abs=1e-6)

    def test_a_collapsing_rate_raises_the_return(self, rich_deal):
        """The 2007 case in miniature: sign the deal at 5.5% and watch the
        coupon fall to 1.5% as the central bank cuts."""
        zirp = _path(rich_deal, [0.055, 0.045, 0.025, 0.015, 0.015])
        assert sponsor_irr(run_lbo(zirp)) > sponsor_irr(run_lbo(rich_deal))

    def test_the_revolver_takes_a_path_too(self, rich_deal):
        d = rich_deal.model_copy(deep=True)
        d.revolver.cash_rate = [0.06, 0.05, 0.04, 0.03, 0.02]
        deal = Assumptions.model_validate(d.model_dump())
        assert deal.rate_schedule(None) == [0.06, 0.05, 0.04, 0.03, 0.02]
        run_lbo(deal)


class TestItIsValidated:
    def test_a_path_of_the_wrong_length_is_rejected(self, rich_deal):
        with pytest.raises(ValueError, match="entries but hold period"):
            _path(rich_deal, [0.05, 0.05])

    def test_a_rate_outside_the_bounds_is_rejected(self, rich_deal):
        """Pydantic checks ge/lt on a bare float; a list slips past that, so the
        bounds are asserted rather than trusted. A coupon of 55 instead of 0.055
        is the commonest input error there is."""
        with pytest.raises(ValueError, match="outside 0-100%"):
            _path(rich_deal, [0.05, 0.05, 55.0, 0.05, 0.05])

    def test_the_same_bound_holds_for_the_revolver(self, rich_deal):
        d = rich_deal.model_copy(deep=True)
        d.revolver.cash_rate = [0.05, 0.05, 12.0, 0.05, 0.05]
        with pytest.raises(ValueError, match="outside 0-100%"):
            Assumptions.model_validate(d.model_dump())


class TestItInteractsCorrectly:
    def test_the_refinancing_spread_still_stacks_on_top(self, rich_deal):
        """Two separate things: the market rate for the year, and the step-up
        for having had to roll. They add."""
        d = _path(rich_deal, [0.05, 0.05, 0.05, 0.05, 0.05])
        d.tranches[0].maturity_years = 2
        d.tranches[0].refinancing_spread = 0.03
        deal = Assumptions.model_validate(d.model_dump())

        r = run_lbo(deal)
        name = deal.tranches[0].name
        t = r.years[2].tranches[name]  # year AFTER the roll
        basis = 0.5 * (t.opening + t.closing)
        assert t.cash_interest == pytest.approx((0.05 + 0.03) * basis, abs=1e-6)

    def test_a_pik_election_uses_the_year_rate(self, rich_deal):
        """Electing moves that year's coupon into PIK — this year's, not the
        first year's."""
        d = _path(rich_deal, [0.05, 0.05, 0.05, 0.05, 0.05])
        d.tranches[-1].pik_toggle = True
        d.operating.ebitda_margin = [0.192, 0.115, 0.11, 0.115, 0.125]
        d.revolver.commitment = 60.0
        deal = Assumptions.model_validate(d.model_dump())
        r = run_lbo(deal)
        assert any(y.pik_elections for y in r.years)


class TestTheHiltonSensitivity:
    """The number the review quoted, reproduced. It is the argument for why the
    limitation had to be re-filed as understating rather than neutral."""

    def test_two_hundred_basis_points_is_worth_a_third_of_a_turn(self):
        from api.case_studies import CASES

        hilton = next(c for c in CASES if c.slug.startswith("hilton")).realised
        base = run_lbo(hilton).moic

        cut = hilton.model_copy(deep=True)
        for tranche in cut.tranches:
            tranche.cash_rate = max(tranche.cash_rate - 0.02, 0.0)
        cut.revolver.cash_rate = max(cut.revolver.cash_rate - 0.02, 0.0)
        lower = run_lbo(Assumptions.model_validate(cut.model_dump())).moic

        assert lower - base > 0.3, (
            f"200bp moved MOIC from {base:.2f}x to {lower:.2f}x — if this stops "
            "being material, the limitations text should stop saying it is"
        )


class TestTheWorkbookCarriesThePath:
    """A rate path that exists in the engine and not in the export would be a
    formula model quietly disagreeing with the model it came from — the exact
    drift the recalculation test exists to prevent."""

    def _acyclic_with_path(self, deal, rates):
        d = deal.model_copy(deep=True)
        d.interest_on_average_balance = False
        d.tranches[0].cash_rate = list(rates)
        return Assumptions.model_validate(d.model_dump())

    def test_it_exports_and_recalculates_to_the_engine(self, rich_deal, tmp_path):
        formulas = pytest.importorskip("formulas")
        openpyxl = pytest.importorskip("openpyxl")
        from lbo_engine.workbook import build_workbook

        rates = [0.02, 0.04, 0.06, 0.08, 0.10]
        deal = self._acyclic_with_path(rich_deal, rates)
        path = str(tmp_path / "rates.xlsx")
        build_workbook(deal).save(path)

        solution = formulas.ExcelModel().loads(path).finish().calculate()
        book = openpyxl.load_workbook(path)

        def value(sheet, cell):
            v = solution[f"'[rates.xlsx]{sheet.upper()}'!{cell}"]
            try:
                return float(v.value[0, 0])
            except Exception:
                return float(v)

        rows = [r[0].row for r in book["Model"].iter_rows(min_col=1, max_col=1)
                if r[0].value == "Cash interest"]
        result = run_lbo(deal)
        name = deal.tranches[0].name
        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            # The first "Cash interest" row is the senior tranche's own.
            assert value("Model", f"{c}{rows[0]}") == pytest.approx(
                year.tranches[name].cash_interest, abs=1e-4), f"year {year.year}"

    def test_it_survives_the_round_trip(self, rich_deal, tmp_path):
        pytest.importorskip("openpyxl")
        from lbo_engine.workbook import build_workbook
        from lbo_engine.workbook_read import read_workbook

        rates = [0.02, 0.04, 0.06, 0.08, 0.10]
        deal = self._acyclic_with_path(rich_deal, rates)
        path = str(tmp_path / "roundtrip.xlsx")
        build_workbook(deal).save(path)

        back = read_workbook(path)
        assert back.tranches[0].cash_rate == pytest.approx(rates)

    def test_a_flat_coupon_stays_a_single_cell(self, rich_deal, tmp_path):
        """Not every deal needs a row per coupon. A flat rate exports as one
        blue cell, because a five-column row of identical numbers is noise."""
        pytest.importorskip("openpyxl")
        import openpyxl

        from lbo_engine.workbook import build_workbook

        path = str(tmp_path / "flat.xlsx")
        build_workbook(rich_deal).save(path)
        wb = openpyxl.load_workbook(path)
        _, ref = next(iter(wb.defined_names["T1_Rate"].destinations))
        assert ":" not in ref, f"a flat coupon exported as a range: {ref}"
