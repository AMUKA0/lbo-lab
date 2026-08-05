"""The Excel export.

The point of this file is one test: open the exported workbook, recalculate it
with an independent formula engine, and assert it agrees with the Python model.
Without that, a formula-driven export is a second implementation of the maths
that can drift from the first, silently, and be wrong in a spreadsheet someone
is pasting into an IC memo.

The recalculation runs against the **opening-balance** convention, because that
is acyclic and can be evaluated outside Excel. The average-balance convention is
genuinely circular — Excel resolves it by iteration and `formulas` reports
`#CIRC!` rather than iterating — so for that variant the tests below assert the
formula shape and the iteration flag instead. Every other line of the model,
including the whole waterfall, the tax build and the exit, is verified numerically.
"""

import os
import tempfile

import pytest

from lbo_engine import Assumptions, InterestLimitation, run_lbo
from lbo_engine.returns import sponsor_irr
from lbo_engine.workbook import build_workbook

openpyxl = pytest.importorskip("openpyxl")


def _acyclic(deal: Assumptions) -> Assumptions:
    payload = deal.model_dump()
    payload["interest_on_average_balance"] = False
    return Assumptions.model_validate(payload)


def _write(deal: Assumptions, name: str) -> str:
    path = os.path.join(tempfile.mkdtemp(), name)
    build_workbook(deal).save(path)
    return path


def _recalculate(path: str):
    """Evaluate the workbook with `formulas`, and return a value getter."""
    formulas = pytest.importorskip("formulas")
    from openpyxl import load_workbook

    solution = formulas.ExcelModel().loads(path).finish().calculate()
    wb = load_workbook(path)
    base = os.path.basename(path)

    def value(sheet: str, cell: str) -> float:
        v = solution[f"'[{base}]{sheet.upper()}'!{cell}"]
        try:
            return float(v.value[0, 0])
        except Exception:
            return float(v)

    def row(sheet: str, label: str) -> int:
        """The LAST row carrying this label — per-tranche blocks reuse names
        like "Cash interest", and the summary line is always below them."""
        hits = [r[0].row for r in wb[sheet].iter_rows(min_col=1, max_col=1)
                if r[0].value == label]
        assert hits, f"no row labelled {label!r} on {sheet}"
        return hits[-1]

    return value, row, wb


class TestItIsTheSameModel:
    def test_every_line_agrees_with_the_engine(self, rich_deal):
        """The load-bearing test. If this fails, the workbook is lying."""
        deal = _acyclic(rich_deal)
        path = _write(deal, "rich.xlsx")
        value, row, _ = _recalculate(path)
        result = run_lbo(deal)

        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            assert value("Model", f"{c}{row('Model', 'EBITDA')}") == pytest.approx(
                year.ebitda, abs=1e-4), f"year {year.year} EBITDA"
            assert -value("Model", f"{c}{row('Model', 'Cash interest')}") == pytest.approx(
                year.cash_interest_total, abs=1e-4), f"year {year.year} interest"
            assert -value("Model", f"{c}{row('Model', 'Tax')}") == pytest.approx(
                year.taxes, abs=1e-4), f"year {year.year} tax"
            assert value("Model", f"{c}{row('Model', 'Closing cash')}") == pytest.approx(
                year.closing_cash, abs=1e-4), f"year {year.year} cash"
            assert value("Model", f"{c}{row('Model', 'Net debt')}") == pytest.approx(
                year.total_debt_closing - year.closing_cash, abs=1e-4), f"year {year.year} net debt"

    def test_the_returns_agree(self, rich_deal):
        deal = _acyclic(rich_deal)
        value, row, _ = _recalculate(_write(deal, "returns.xlsx"))
        result = run_lbo(deal)

        assert value("S&U", f"B{row('S&U', 'Sponsor equity (the plug)')}") == pytest.approx(
            result.entry_equity, abs=1e-4)
        assert value("Returns", f"B{row('Returns', 'Exit equity')}") == pytest.approx(
            result.exit_equity, abs=1e-4)
        assert value("Returns", f"B{row('Returns', 'MOIC')}") == pytest.approx(
            result.moic, abs=1e-6)

    def test_the_checks_sheet_passes(self, rich_deal):
        """The workbook's own invariants, evaluated. A Checks sheet that is
        never tested is decoration."""
        value, row, wb = _recalculate(_write(_acyclic(rich_deal), "checks.xlsx"))
        labels = [r[0].value for r in wb["Checks"].iter_rows(min_col=1, max_col=1)
                  if r[0].value and r[0].row >= 4]
        assert labels, "no checks written"
        for label in labels:
            r = row("Checks", label)
            kind = wb["Checks"].cell(row=r, column=3).value
            v = value("Checks", f"B{r}")
            # An identity must be zero; a limit only has to be the right side
            # of it. Asserting equality on a headroom check would fail on every
            # healthy deal.
            if kind == "must be 0":
                assert v == pytest.approx(0.0, abs=0.01), f"{label} = {v}"
            else:
                assert v >= -0.01, f"{label} = {v}"


class TestTheInterestLimitation:
    """§163(j) in the workbook. The line above checks tax, which would agree
    even if the cap silently did nothing — so these pin the block itself, and
    that it binds on the fixture rather than sitting inert."""

    def test_the_cap_binds_on_the_fixture(self, rich_deal):
        first = run_lbo(_acyclic(rich_deal)).years[0]
        assert first.interest_cf_closing > 0, (
            "the fixture must actually be capped, or the recalculation below "
            "proves nothing about §163(j)"
        )

    def test_the_carryforward_agrees_with_the_engine(self, rich_deal):
        deal = _acyclic(rich_deal)
        value, row, _ = _recalculate(_write(deal, "capped.xlsx"))
        result = run_lbo(deal)

        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            for label, expected in (
                ("Business interest (§163(j))", year.business_interest),
                ("Deductible capacity", year.interest_capacity),
                ("Interest deducted", year.interest_deducted),
                ("Disallowed interest carried forward", year.interest_cf_closing),
                ("Taxable income", year.taxable_income),
            ):
                assert value("Model", f"{c}{row('Model', label)}") == pytest.approx(
                    expected, abs=1e-4), f"year {year.year} {label}"

    def test_switching_the_cap_off_in_the_sheet_restores_the_deduction(self, rich_deal):
        """The switch is the point of exporting it as an input: an analyst asked
        to run the deal at pre-2018 tax changes one cell, not a formula."""
        path = _write(_acyclic(rich_deal), "uncapped.xlsx")
        wb = openpyxl.load_workbook(path)
        sheet, ref = next(iter(wb.defined_names["Interest_Limit_On"].destinations))
        wb[sheet][ref.replace("$", "")].value = 0
        wb.save(path)

        value, row, _ = _recalculate(path)
        off = run_lbo(_acyclic(rich_deal).model_copy(
            update={"interest_limitation": InterestLimitation(enabled=False)}))
        for i, year in enumerate(off.years):
            c = chr(ord("B") + i)
            assert value("Model", f"{c}{row('Model', 'Interest deducted')}") == pytest.approx(
                year.business_interest, abs=1e-4), f"year {year.year}"
            assert -value("Model", f"{c}{row('Model', 'Tax')}") == pytest.approx(
                year.taxes, abs=1e-4), f"year {year.year} tax"


class TestInterestOnCash:
    def test_the_income_row_agrees_with_the_engine(self, rich_deal):
        """Written into the sheet ABOVE the cash rows it depends on, so this is
        also a check that the forward reference resolves."""
        deal = _acyclic(rich_deal)
        deal.cash_sweep_pct = 0.0  # keep a balance to earn on
        deal.cash_deposit_rate = 0.04
        value, row, _ = _recalculate(_write(deal, "deposits.xlsx"))
        result = run_lbo(deal)

        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            assert year.interest_income > 0
            assert value("Model", f"{c}{row('Model', 'Interest income on cash')}") == (
                pytest.approx(year.interest_income, abs=1e-4)), f"year {year.year}"
            # And it must reach the cap, not just the P&L.
            assert value("Model", f"{c}{row('Model', 'Deductible capacity')}") == (
                pytest.approx(year.interest_capacity, abs=1e-4)), f"year {year.year} cap"


class TestCovenantsAndMaturity:
    def test_a_cov_lite_deal_gets_no_covenant_rows(self, rich_deal):
        """The absence is the fact worth showing. A row reading 'n/a' invites
        someone to fill it in with a number the credit agreement never had."""
        _, _, wb = _recalculate(_write(_acyclic(rich_deal), "covlite.xlsx"))
        labels = [r[0].value for r in wb["Checks"].iter_rows(min_col=1, max_col=1)]
        assert not any("covenant" in (v or "").lower() for v in labels)

    def test_covenant_headroom_recalculates_and_reads_OK(self, rich_deal):
        from lbo_engine import Covenants

        deal = _acyclic(rich_deal)
        deal.covenants = Covenants(
            net_leverage_ceiling=[6.5, 6.0, 5.5, 5.0, 4.5],
            interest_coverage_floor=2.0,
        )
        run_lbo(deal)  # comfortably inside both, so both rows must read OK

        value, row, wb = _recalculate(_write(deal, "covenants.xlsx"))
        for label in ("Leverage covenant headroom, tightest year (turns)",
                      "Coverage covenant headroom, tightest year (turns)"):
            r = row("Checks", label)
            assert value("Checks", f"B{r}") > 0, label

    def test_the_headroom_is_struck_against_each_year_own_ceiling(self, rich_deal):
        """A step-down schedule tests a different level every year. Using one
        number for all of them would be right in exactly one year."""
        from lbo_engine import Covenants

        deal = _acyclic(rich_deal)
        ceilings = [6.5, 6.0, 5.5, 5.0, 4.5]
        deal.covenants = Covenants(net_leverage_ceiling=ceilings)
        value, row, _ = _recalculate(_write(deal, "stepdown.xlsx"))
        result = run_lbo(deal)

        expected = min(
            c - (y.total_debt_closing - y.closing_cash) / y.ebitda
            for c, y in zip(ceilings, result.years)
        )
        r = row("Checks", "Leverage covenant headroom, tightest year (turns)")
        assert value("Checks", f"B{r}") == pytest.approx(expected, abs=1e-6)

    def test_an_unrefinanced_maturity_repays_in_full_in_the_sheet(self, simple_deal):
        deal = _acyclic(simple_deal)
        deal.hold_years = 3
        deal.tranches[0].leverage_turns = 0.2
        deal.tranches[0].maturity_years = 3
        deal.tranches[0].refinance_at_maturity = False

        value, row, _ = _recalculate(_write(deal, "wall.xlsx"))
        result = run_lbo(deal)
        name = deal.tranches[0].name
        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            assert value("Model", f"{c}{row('Model', 'Closing')}") == pytest.approx(
                year.tranches[name].closing, abs=1e-4), f"year {year.year}"
        assert result.years[-1].tranches[name].closing == pytest.approx(0.0, abs=1e-9)


class TestTheCircularity:
    def test_iteration_is_enabled_when_interest_is_circular(self, rich_deal):
        """Without this the workbook opens to a wall of circular-reference
        warnings, and an analyst reasonably concludes it is broken."""
        wb = build_workbook(rich_deal)
        assert rich_deal.interest_on_average_balance
        assert wb.calculation.iterate is True
        assert wb.calculation.iterateCount >= 100
        assert wb.calculation.iterateDelta <= 1e-6

    def test_the_average_balance_formula_is_actually_circular(self, rich_deal):
        """Interest must reference the closing balance it helps determine — that
        is the convention, and losing it would silently change the model."""
        wb = build_workbook(rich_deal)
        ws = wb["Model"]
        interest = [
            ws.cell(row=r[0].row, column=2).value
            for r in ws.iter_rows(min_col=1, max_col=1)
            if r[0].value == "Cash interest"
        ]
        assert any("AVERAGE(" in str(f) for f in interest), interest

    def test_the_breaker_produces_an_acyclic_workbook(self, rich_deal):
        wb = build_workbook(_acyclic(rich_deal))
        assert wb.calculation.iterate is False
        ws = wb["Model"]
        formulas_used = [
            ws.cell(row=r[0].row, column=2).value
            for r in ws.iter_rows(min_col=1, max_col=1)
            if r[0].value == "Cash interest"
        ]
        assert not any("AVERAGE(" in str(f) for f in formulas_used)


class TestItIsReadableByAnalysts:
    def test_calculated_cells_contain_formulas_not_values(self, rich_deal):
        """The whole difference between this and a CSV. A hardcoded number in a
        calculation row cannot be audited or flexed."""
        wb = build_workbook(rich_deal)
        ws = wb["Model"]
        checked = 0
        for r in ws.iter_rows(min_col=1, max_col=1):
            if r[0].value in ("EBITDA", "Net debt", "Closing cash", "Tax"):
                for i in range(rich_deal.hold_years):
                    v = ws.cell(row=r[0].row, column=2 + i).value
                    assert isinstance(v, str) and v.startswith("="), f"{r[0].value} is hardcoded"
                    checked += 1
        assert checked > 0

    def test_inputs_are_named_ranges(self, rich_deal):
        """Named ranges survive an analyst inserting a row. Cell addresses do not."""
        wb = build_workbook(rich_deal)
        names = set(wb.defined_names)
        for expected in ("Entry_EBITDA", "Entry_Multiple", "Exit_Multiple",
                         "Tax_Rate", "Sweep_Pct", "Minimum_Cash"):
            assert expected in names, expected

    def test_inputs_are_blue_and_formulas_are_not(self, rich_deal):
        """Blue means "type here". Getting this wrong invites someone to
        overwrite a formula."""
        wb = build_workbook(rich_deal)
        ebitda = next(r[0].row for r in wb["Inputs"].iter_rows(min_col=1, max_col=1)
                      if r[0].value == "Entry EBITDA ($m)")
        assert wb["Inputs"].cell(row=ebitda, column=2).font.color.rgb.endswith("0000CC")

        model = wb["Model"]
        row = next(r[0].row for r in model.iter_rows(min_col=1, max_col=1)
                   if r[0].value == "EBITDA")
        assert not model.cell(row=row, column=2).font.color.rgb.endswith("0000CC")

    def test_the_sheets_are_separated_by_role(self, rich_deal):
        assert build_workbook(rich_deal).sheetnames == [
            "Inputs", "S&U", "Model", "Returns", "Checks"]


class TestMidHoldCapitalEvents:
    """The workbook used to refuse these, on the grounds that they were
    decisions the engine makes by search rather than by formula. That was true
    of exactly one of them. Divestitures and sponsor support are fully specified
    inputs; a recap sized by target leverage is MAX(target x EBITDA - net debt,
    0), which is arithmetic. Only the PIK election is a search, and exporting
    the engine's choice as an overridable input settles it.
    """

    def _with(self, deal: Assumptions, **events) -> Assumptions:
        payload = deal.model_dump()
        payload["interest_on_average_balance"] = False
        payload.update(events)
        return Assumptions.model_validate(payload)

    def _worst_disagreement(self, deal: Assumptions) -> float:
        value, row, _ = _recalculate(_write(deal, "ev.xlsx"))
        result = run_lbo(deal)
        worst = 0.0
        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            for excel, engine in (
                (value("Model", f"{c}{row('Model', 'Net debt')}"),
                 year.total_debt_closing - year.closing_cash),
                (value("Model", f"{c}{row('Model', 'Closing cash')}"), year.closing_cash),
                (-value("Model", f"{c}{row('Model', 'Cash interest')}"), year.cash_interest_total),
            ):
                worst = max(worst, abs(excel - engine))
        return worst

    DIVEST = [{"year": 3, "proceeds": 80.0, "fee_pct": 0.01, "taxable_gain": 20.0,
               "revenue_removed": 30.0, "label": "A unit"}]
    RECAP = [{"year": 4, "amount": 50.0, "target_leverage_turns": None,
              "tranche": None, "financing_fee_pct": 0.02}]
    INJECT = [{"year": 2, "amount": 40.0, "debt_retired": 60.0, "label": "Rescue"}]

    def test_a_divestiture_agrees_with_the_engine(self, rich_deal):
        assert self._worst_disagreement(
            self._with(rich_deal, divestitures=self.DIVEST)) == pytest.approx(0, abs=1e-6)

    def test_a_recap_agrees_with_the_engine(self, rich_deal):
        assert self._worst_disagreement(
            self._with(rich_deal, recaps=self.RECAP)) == pytest.approx(0, abs=1e-6)

    def test_sponsor_support_agrees_with_the_engine(self, rich_deal):
        assert self._worst_disagreement(
            self._with(rich_deal, injections=self.INJECT)) == pytest.approx(0, abs=1e-6)

    def test_all_three_together_agree_with_the_engine(self, rich_deal):
        """Events interact — a divestiture changes the balance a later recap is
        sized against — so agreeing separately is not the same as agreeing at
        once."""
        assert self._worst_disagreement(self._with(
            rich_deal, divestitures=self.DIVEST, recaps=self.RECAP,
            injections=self.INJECT)) == pytest.approx(0, abs=1e-6)

    def test_divested_revenue_does_not_release_working_capital(self, rich_deal):
        """The subtle one, and the reason the divestiture case disagreed at
        first: revenue that leaves with a sold business must not drive a
        working-capital release, because that cash went with the business."""
        wb = build_workbook(self._with(rich_deal, divestitures=self.DIVEST))
        assert "Divest_Revenue" in set(wb.defined_names)
        ws = wb["Model"]
        row = next(r[0].row for r in ws.iter_rows(min_col=1, max_col=1)
                   if r[0].value == "Change in working capital")
        # Year two onward must net off the prior year's departed revenue.
        assert "Divest_Revenue" not in str(ws.cell(row=row, column=2).value)
        assert any("$" in str(ws.cell(row=row, column=2 + i).value)
                   for i in range(1, rich_deal.hold_years))

    def test_recap_debt_joins_the_amortising_base(self, rich_deal):
        """Otherwise the incremental debt enjoys a permanent amortisation
        holiday — the same bug the engine had until a reviewer found it."""
        wb = build_workbook(self._with(rich_deal, recaps=self.RECAP))
        ws = wb["Model"]
        row = next(r[0].row for r in ws.iter_rows(min_col=1, max_col=1)
                   if r[0].value == "Mandatory amortisation")
        later = str(ws.cell(row=row, column=2 + rich_deal.hold_years - 1).value)
        assert later.count("+") >= 1, later

    def test_a_pik_toggle_election_is_exported_as_an_overridable_input(self, rich_deal):
        """The one genuine search. The engine's decision becomes a blue 1/0 the
        analyst can change, rather than a refusal or a hidden assumption."""
        payload = rich_deal.model_dump()
        payload["tranches"][-1]["pik_toggle"] = True
        wb = build_workbook(Assumptions.model_validate(payload))
        assert f"T{len(rich_deal.tranches)}_Elected" in set(wb.defined_names)

    def test_nothing_is_refused_any_more(self, rich_deal):
        """The boundary is gone. Kept as a test so its removal is deliberate."""
        build_workbook(self._with(
            rich_deal, divestitures=self.DIVEST, recaps=self.RECAP, injections=self.INJECT))


class TestItPrints:
    """An IC exhibit gets printed. The absence of page setup is loud: the first
    time someone hits Ctrl-P and gets one column alone on page four, the file
    stops being credible regardless of what the numbers say."""

    def test_the_schedule_is_landscape_and_fits_the_page_width(self, rich_deal):
        wb = build_workbook(rich_deal)
        model = wb["Model"]
        assert model.page_setup.orientation == "landscape"
        assert model.page_setup.fitToWidth == 1
        # Zero means "as many pages as it takes" downwards. Fitting to one page
        # vertically would shrink a long schedule until nobody could read it,
        # which is the classic way this gets done badly.
        assert model.page_setup.fitToHeight == 0
        assert model.sheet_properties.pageSetUpPr.fitToPage is True

    def test_a_short_schedule_stays_portrait(self, simple_deal):
        assert build_workbook(simple_deal)["Model"].page_setup.orientation == "portrait"

    def test_the_year_headers_repeat_on_every_page(self, rich_deal):
        """A schedule running past one page is unreadable without them, and
        scrolling back is not an option on paper."""
        # openpyxl normalises this to absolute row references on write.
        assert build_workbook(rich_deal)["Model"].print_title_rows == "$1:$3"

    def test_every_sheet_has_a_print_area_and_a_footer(self, rich_deal):
        wb = build_workbook(rich_deal)
        for sheet in wb.worksheets:
            assert sheet.print_area, f"{sheet.title} has no print area"
            # An explicit area stops a stray value in a far column — an
            # analyst's scratch note, which is exactly what people do — from
            # silently dragging a blank page into the printout. openpyxl
            # qualifies the ref with the sheet name on write.
            assert "$A$1:" in str(sheet.print_area), sheet.print_area
            assert sheet.title in (sheet.oddFooter.left.text or "")
            assert "Page" in (sheet.oddFooter.right.text or "")

    def test_a_partial_workbook_prints_too(self, rich_deal):
        """The sheet set differs on a partial export — Returns is rebuilt — so
        the setup has to run after that, not before."""
        broken = rich_deal.model_copy(deep=True)
        broken.operating.ebitda_margin = [0.192, 0.19, 0.03, 0.03, 0.03]
        broken.revolver.commitment = 20.0
        from lbo_engine.workbook import build_partial_workbook

        wb = build_partial_workbook(broken)
        for sheet in wb.worksheets:
            assert sheet.print_area, f"{sheet.title} has no print area"


class TestTheReturnsSheetSeesEveryFlow:
    """The bug this class exists for: MOIC was exit equity over the closing
    cheque, and IRR was that multiple compounded over the hold. Both are correct
    ONLY when nothing happens between close and exit — and the sheet asserted
    the equivalence in a printed comment on files where it was false.

    On HCA, whose sponsors took a $4.3bn recap dividend in year four, the export
    reported 2.27× against the engine's 2.95×. The debt that funded the dividend
    was in the model; the dividend was not. On Hilton the same omission ran the
    other way, flattering the multiple by ignoring $800m of rescue capital.

    Every check on both files read OK, because the four-line Excel bridge was
    algebraically tautological with the wrong entry equity.
    """

    def test_moic_and_irr_agree_with_the_engine_on_a_deal_with_events(
        self, eventful_deal
    ):
        deal = _acyclic(eventful_deal)
        value, row, _ = _recalculate(_write(deal, "events.xlsx"))
        result = run_lbo(deal)

        assert deal.recaps and deal.injections, "the fixture must exercise both"
        assert value("Returns", f"B{row('Returns', 'MOIC')}") == pytest.approx(
            result.moic, abs=1e-4)
        assert value("Returns", f"B{row('Returns', 'IRR')}") == pytest.approx(
            sponsor_irr(result), abs=1e-4)

    def test_the_flow_vector_carries_the_dividend_and_the_injection(
        self, eventful_deal
    ):
        """Not just the totals — the money has to appear in the year it moved,
        or the IRR is right by luck."""
        deal = _acyclic(eventful_deal)
        value, row, _ = _recalculate(_write(deal, "flows.xlsx"))
        result = run_lbo(deal)

        flows = result.equity_cash_flows
        net = row("Returns", "Net cash flow")
        for i, expected in enumerate(flows):
            column = chr(ord("B") + i)
            assert value("Returns", f"{column}{net}") == pytest.approx(
                expected, abs=1e-4), f"year {i} flow"

    def test_total_invested_includes_rescue_capital(self, eventful_deal):
        deal = _acyclic(eventful_deal)
        value, row, _ = _recalculate(_write(deal, "invested.xlsx"))
        result = run_lbo(deal)

        assert value("Returns", f"B{row('Returns', 'Total invested')}") == pytest.approx(
            result.total_invested, abs=1e-4)
        assert result.total_invested > result.entry_equity, (
            "the fixture must inject something, or this proves nothing"
        )

    def test_the_bridge_carries_the_event_lines(self, eventful_deal):
        """`returns.py` splits divestiture proceeds out of deleveraging and
        carries follow-on equity as a negative line, precisely so a sale is not
        reported as operational paydown. The Excel bridge dropped both."""
        _, _, wb = _recalculate(_write(_acyclic(eventful_deal), "bridge.xlsx"))
        labels = [r[0].value for r in wb["Returns"].iter_rows(min_col=1, max_col=1)
                  if r[0].value]
        for expected in ("Divestiture proceeds", "Recapitalisation", "Follow-on equity"):
            assert expected in labels, f"{expected} missing from the Excel bridge"

    def test_the_bridge_still_reconciles_with_events(self, eventful_deal):
        value, row, wb = _recalculate(_write(_acyclic(eventful_deal), "recon.xlsx"))
        r = row("Checks", "Bridge reconciles to the equity gain")
        assert value("Checks", f"B{r}") == pytest.approx(0.0, abs=0.01)

    def test_a_deal_with_no_events_is_unchanged(self, rich_deal):
        """The quiet path must not move. A fix that shifts every existing
        schedule is a different bug."""
        deal = _acyclic(rich_deal)
        value, row, _ = _recalculate(_write(deal, "quiet.xlsx"))
        result = run_lbo(deal)
        assert value("Returns", f"B{row('Returns', 'MOIC')}") == pytest.approx(
            result.moic, abs=1e-6)


class TestEveryBlueCellDrivesSomething:
    """A blue input that changes no number is the commonest cause of a wrong
    answer in a flexed model, and this sheet's own header says "blue cells are
    inputs". `T3_Sweepable` was blue, labelled, named — and referenced by no
    formula, because the export baked the flag in at write time."""

    def _sweeping_deal(self, rich_deal):
        """A structure where the junior CAN sweep: the senior is small enough to
        be repaid, so there is a pool left for the tranche below it. On the
        stock fixture the senior never clears and the flag is unobservable —
        which is a fact about the waterfall, not about the flag."""
        d = _acyclic(rich_deal).model_copy(deep=True)
        d.tranches[0].leverage_turns = 0.5
        return Assumptions.model_validate(d.model_dump())

    def test_the_formula_reads_the_named_input(self, rich_deal):
        """Structural: every sweep row must consult its own blue cell, whatever
        the waterfall happens to allow on a given deal."""
        wb = build_workbook(rich_deal)
        rows = [r[0].row for r in wb["Model"].iter_rows(min_col=1, max_col=1)
                if r[0].value == "Cash sweep"]
        assert len(rows) == len(rich_deal.tranches)
        for index, row in enumerate(rows, start=1):
            formula = wb["Model"].cell(row=row, column=2).value
            assert f"T{index}_Sweepable" in str(formula), (
                f"tranche {index}'s sweep ignores its own flag: {formula}"
            )

    def test_flipping_sweepable_actually_changes_the_sweep(self, rich_deal):
        deal = self._sweeping_deal(rich_deal)
        assert not deal.tranches[1].sweepable, "fixture must start non-sweeping"

        path = _write(deal, "sweepflag.xlsx")
        before, row, _ = _recalculate(path)
        junior = [r[0].row for r in openpyxl.load_workbook(path)["Model"].iter_rows(
            min_col=1, max_col=1) if r[0].value == "Cash sweep"][-1]
        assert before("Model", f"C{junior}") == pytest.approx(0.0, abs=1e-9)

        wb = openpyxl.load_workbook(path)
        sheet, ref = next(iter(wb.defined_names["T2_Sweepable"].destinations))
        wb[sheet][ref.replace("$", "")].value = 1
        wb.save(path)

        after, _, _ = _recalculate(path)
        assert abs(after("Model", f"C{junior}")) > 1.0, (
            "flipping the blue Sweepable cell changed nothing — it is dead again"
        )

    def test_the_flipped_flag_matches_the_engine(self, rich_deal):
        """And the number it produces is the engine's, not merely non-zero."""
        swept = self._sweeping_deal(rich_deal).model_copy(deep=True)
        swept.tranches[1].sweepable = True
        swept = Assumptions.model_validate(swept.model_dump())

        path = _write(swept, "sweepon.xlsx")
        value, _, wb = _recalculate(path)
        junior = [r[0].row for r in wb["Model"].iter_rows(min_col=1, max_col=1)
                  if r[0].value == "Cash sweep"][-1]

        result = run_lbo(swept)
        name = swept.tranches[1].name
        for i, year in enumerate(result.years):
            c = chr(ord("B") + i)
            assert -value("Model", f"{c}{junior}") == pytest.approx(
                year.tranches[name].sweep_repayment, abs=1e-4), f"year {year.year}"


class TestTheRoundTripKeepsTheEvents:
    """Exporting a deal and uploading it back unchanged must return the SAME
    deal. It did not: the reader had no notion of recaps, injections or
    divestitures, so HCA came back with its $4.3bn recap silently gone and
    Hilton came back as an unfinanceable structure — the site rejecting its own
    artefact, with no warning on either."""

    def _round_trip(self, deal, name, tmp_path):
        from lbo_engine.workbook_read import read_workbook

        path = str(tmp_path / f"{name}.xlsx")
        build_workbook(deal).save(path)
        return read_workbook(path)

    def test_a_recap_survives(self, eventful_deal, tmp_path):
        back = self._round_trip(eventful_deal, "recap", tmp_path)
        assert len(back.recaps) == len(eventful_deal.recaps)
        assert back.recaps[0].year == eventful_deal.recaps[0].year

    def test_an_injection_survives(self, eventful_deal, tmp_path):
        back = self._round_trip(eventful_deal, "inject", tmp_path)
        assert len(back.injections) == len(eventful_deal.injections)
        assert back.injections[0].amount == pytest.approx(
            eventful_deal.injections[0].amount)

    def test_a_divestiture_survives(self, eventful_deal, tmp_path):
        back = self._round_trip(eventful_deal, "divest", tmp_path)
        assert len(back.divestitures) == len(eventful_deal.divestitures)

    def test_the_returns_are_unchanged_by_the_trip(self, eventful_deal, tmp_path):
        """The test that actually matters. Field counts can match while the
        deal has quietly become a different one."""
        back = self._round_trip(eventful_deal, "same", tmp_path)
        original, returned = run_lbo(eventful_deal), run_lbo(back)
        assert returned.moic == pytest.approx(original.moic, rel=1e-6)
        assert sponsor_irr(returned) == pytest.approx(
            sponsor_irr(original), rel=1e-6)

    def test_a_deal_with_no_events_still_round_trips_clean(self, rich_deal, tmp_path):
        """An empty event row must read as "nothing happened", not as an event
        of zero — a DividendRecap(amount=0) would be rejected outright."""
        back = self._round_trip(rich_deal, "quiet", tmp_path)
        assert back.recaps == [] and back.injections == [] and back.divestitures == []


class TestTheFileRendersWithoutBeingCalculated:
    """openpyxl writes `<f>=…</f>` and nothing else, because it cannot
    calculate. Excel fills those in on open — `fullCalcOnLoad` is set — but
    anything that RENDERS the file rather than calculating it shows every
    calculated cell blank: Explorer and Quick Look previews, Gmail, Drive,
    Slack, Teams, GitHub. On the Model sheet that was 259 of 326 cells.

    So the file a recruiter previews before opening it looked broken, which is
    the worst possible first impression for the artefact this project calls its
    strongest. The answers now ship alongside the formulas.
    """

    def _cells(self, payload: bytes):
        """Every formula cell in the archive, with whether it carries a value."""
        import io as _io
        import re
        import zipfile

        found = []
        with zipfile.ZipFile(_io.BytesIO(payload)) as z:
            for name in z.namelist():
                if not name.startswith("xl/worksheets/sheet"):
                    continue
                xml = z.read(name).decode("utf-8", "ignore")
                for ref, body in re.findall(
                    r'<c r="([A-Z]+\d+)"[^>]*>((?:(?!</c>).)*?)</c>', xml, re.S
                ):
                    if "<f" in body:
                        found.append((name, ref, "<v>" in body))
        return found

    def test_every_formula_cell_carries_a_cached_result(self, rich_deal):
        from lbo_engine.workbook import workbook_bytes

        cells = self._cells(workbook_bytes(rich_deal))
        assert cells, "no formula cells at all — the export is not a live model"
        blank = [f"{n.split('/')[-1]}!{r}" for n, r, cached in cells if not cached]
        assert not blank, (
            f"{len(blank)} of {len(cells)} formula cells would render blank in a "
            f"preview: {blank[:8]}"
        )

    def test_it_holds_on_a_deal_with_every_mechanic(self, eventful_deal):
        """Recaps, injections and divestitures each add rows, and a row added
        without a cached value is a blank line in the middle of a schedule."""
        from lbo_engine.workbook import workbook_bytes

        blank = [r for _, r, cached in self._cells(workbook_bytes(eventful_deal))
                 if not cached]
        assert not blank, blank[:8]

    def test_the_cached_values_agree_with_their_own_formulas(self, rich_deal):
        """The risk this creates, closed. A cached value that disagrees with the
        formula beside it is worse than a blank cell: Excel would show one
        number and recalculate to another.

        Checked on the acyclic convention, which an independent evaluator can
        reproduce — the circular variant is the reason the values come from the
        engine in the first place.
        """
        import io as _io

        from openpyxl import load_workbook

        from lbo_engine.workbook import workbook_bytes

        deal = _acyclic(rich_deal)
        payload = workbook_bytes(deal)
        path = os.path.join(tempfile.mkdtemp(), "cached.xlsx")
        with open(path, "wb") as fh:
            fh.write(payload)

        value, _, _ = _recalculate(path)
        # `load_workbook` without data_only gives the formulas; with it, the
        # cached values we wrote.
        cached = load_workbook(path, data_only=True)

        compared = 0
        for sheet in ("Model", "Returns", "S&U"):
            for row in cached[sheet].iter_rows():
                for cell in row:
                    if not isinstance(cell.value, (int, float)):
                        continue
                    try:
                        computed = value(sheet, cell.coordinate)
                    except Exception:
                        continue  # not a formula cell, or unresolved
                    assert cell.value == pytest.approx(computed, rel=1e-6, abs=1e-4), (
                        f"{sheet}!{cell.coordinate}: cached {cell.value} but the "
                        f"formula computes {computed}"
                    )
                    compared += 1
        assert compared > 100, f"only {compared} cells compared — the test is hollow"

    def test_the_check_verdicts_render_too(self, rich_deal):
        """The Result column is what a reader looks for first, and it holds text
        rather than a number — which needs `t="str"` on the cell or Excel reads
        the cached value as a shared-string index and shows an unrelated word."""
        import io as _io

        from openpyxl import load_workbook

        from lbo_engine.workbook import workbook_bytes

        wb = load_workbook(_io.BytesIO(workbook_bytes(rich_deal)), data_only=True)
        verdicts = [wb["Checks"].cell(row=r, column=4).value
                    for r in range(4, wb["Checks"].max_row + 1)]
        assert verdicts and all(v == "OK" for v in verdicts), verdicts

    def test_the_test_column_is_not_itself_a_formula(self, rich_deal):
        """It used to read "= 0", which Excel treats as a formula and evaluates
        to 0 — so the column meant to say what each check requires displayed a
        bare zero instead."""
        wb = build_workbook(rich_deal)
        for r in range(4, wb["Checks"].max_row + 1):
            kind = wb["Checks"].cell(row=r, column=3).value
            if kind:
                assert not str(kind).startswith("="), kind
                assert "must be" in str(kind), kind
